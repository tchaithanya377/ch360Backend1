from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import logout
from django.contrib import messages
from django.db import connection, models
from django.apps import apps
from django.http import JsonResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from accounts.models import User, Role, Permission, AuthIdentifier, UserSession, AuditLog, FailedLogin
from students.models import Student, StudentEnrollmentHistory, StudentDocument, CustomField, StudentImport
from departments.models import Department
from academics.models import AcademicProgram
from faculty.models import Faculty, FacultySubject, FacultySchedule, FacultyLeave, FacultyPerformance, FacultyDocument, CustomField as FacultyCustomField, CustomFieldValue
from django.utils import timezone
from django.db.models import Avg
from django.http import HttpResponse
from django.db.models import Q, Count
import csv
import os

try:
    from .models import APICollection, APIEnvironment, APIRequest, APITest, APITestResult, APITestSuite, APITestSuiteResult, APIAutomation
except Exception:  # Models were removed; provide dummies for view compatibility
    APICollection = APIEnvironment = APIRequest = APITest = APITestResult = APITestSuite = APITestSuiteResult = APIAutomation = None
from academics.models import Course, Syllabus, Timetable, CourseEnrollment, AcademicCalendar, AcademicProgram, CourseSection
from departments.models import Department
from attendance.models import AttendanceSession, AttendanceRecord
from enrollment.models import EnrollmentRule, CourseAssignment, FacultyAssignment, StudentEnrollmentPlan, PlannedCourse, EnrollmentRequest, WaitlistEntry
from grads.models import GradeScale, Term, CourseResult, TermGPA, GraduateRecord
from rnd.models import Researcher as RndResearcher, Grant as RndGrant, Project as RndProject, Publication as RndPublication, Patent as RndPatent, Dataset as RndDataset, Collaboration as RndCollaboration
from fees.models import FeeCategory, FeeStructure, FeeStructureDetail, StudentFee, Payment, FeeWaiver, FeeDiscount, FeeReceipt
from transportation.models import Vehicle, Driver, Route, Stop, RouteStop, VehicleAssignment, TripSchedule, TransportPass
from transportation.forms import (
    VehicleForm,
    DriverForm,
    RouteForm,
    StopForm,
    RouteStopForm,
    VehicleAssignmentForm,
    TripScheduleForm,
    TransportPassForm,
)
from django.contrib import messages
from mentoring.models import Mentorship, Project, Meeting, Feedback
from feedback.models import Feedback as UnivFeedback, FeedbackCategory as UnivFeedbackCategory, FeedbackTag as UnivFeedbackTag, FeedbackComment as UnivFeedbackComment, FeedbackAttachment as UnivFeedbackAttachment, FeedbackVote as UnivFeedbackVote
from assignments.models import (
    Assignment, AssignmentSubmission, AssignmentFile, AssignmentGrade, 
    AssignmentComment, AssignmentCategory, AssignmentGroup, AssignmentTemplate
)
from assignments.forms import AssignmentForm

# Ensure is_admin is defined before any decorators use it
def is_admin(user):
    return user.is_authenticated and user.is_staff
@login_required
@user_passes_test(is_admin)
def feedback_dashboard(request):
    stats = {
        'total': UnivFeedback.objects.count(),
        'open': UnivFeedback.objects.filter(status='open').count(),
        'in_review': UnivFeedback.objects.filter(status='in_review').count(),
        'resolved': UnivFeedback.objects.filter(status='resolved').count(),
        'closed': UnivFeedback.objects.filter(status='closed').count(),
    }
    recent = UnivFeedback.objects.select_related('category', 'created_by').order_by('-created_at')[:10]
    return render(request, 'dashboard/feedback/dashboard.html', {
        'stats': stats,
        'recent': recent,
        'categories': UnivFeedbackCategory.objects.filter(is_active=True).order_by('name'),
        'tags': UnivFeedbackTag.objects.order_by('name'),
    })


@login_required
@user_passes_test(is_admin)
def feedback_items_list(request):
    items = UnivFeedback.objects.select_related('category', 'created_by', 'department', 'course', 'section', 'faculty', 'syllabus').prefetch_related('tags').all()
    # Filters
    q = request.GET.get('q')
    status_f = request.GET.get('status')
    category_f = request.GET.get('category')
    department_f = request.GET.get('department')
    course_f = request.GET.get('course')
    faculty_f = request.GET.get('faculty')
    if q:
        items = items.filter(Q(title__icontains=q) | Q(description__icontains=q))
    if status_f:
        items = items.filter(status=status_f)
    if category_f:
        items = items.filter(category_id=category_f)
    if department_f:
        items = items.filter(department_id=department_f)
    if course_f:
        items = items.filter(course_id=course_f)
    if faculty_f:
        items = items.filter(faculty_id=faculty_f)
    items = items.order_by('-created_at')
    return render(request, 'dashboard/feedback/items_list.html', {
        'items': items,
        'categories': UnivFeedbackCategory.objects.filter(is_active=True).order_by('name'),
        'departments': Department.objects.order_by('name'),
        'courses': Course.objects.order_by('code'),
        'faculties': Faculty.objects.order_by('-created_at')[:100],
        'status_choices': UnivFeedback.Status.choices,
        'current': {
            'q': q, 'status': status_f, 'category': category_f, 'department': department_f, 'course': course_f, 'faculty': faculty_f
        }
    })


@login_required
@user_passes_test(is_admin)
def feedback_item_detail(request, item_id):
    item = get_object_or_404(UnivFeedback.objects.select_related('category', 'created_by', 'department', 'course', 'section', 'faculty', 'syllabus'), id=item_id)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'comment':
            content = request.POST.get('content', '').strip()
            if content:
                UnivFeedbackComment.objects.create(feedback=item, commented_by=request.user, content=content, is_internal=bool(request.POST.get('is_internal')))
                messages.success(request, 'Comment added')
            return redirect('dashboard:feedback_item_detail', item_id=item.id)
        if action == 'vote':
            is_up = request.POST.get('is_up') == '1'
            UnivFeedbackVote.objects.update_or_create(feedback=item, voted_by=request.user, defaults={'is_upvote': is_up})
            messages.success(request, 'Vote recorded')
            return redirect('dashboard:feedback_item_detail', item_id=item.id)
        if action == 'update_status':
            new_status = request.POST.get('status')
            if new_status in dict(UnivFeedback.Status.choices):
                item.status = new_status
                item.save(update_fields=['status'])
                messages.success(request, 'Status updated')
            return redirect('dashboard:feedback_item_detail', item_id=item.id)
    comments = item.comments.select_related('commented_by').order_by('-created_at')
    votes = item.votes.select_related('voted_by')
    score = votes.filter(is_upvote=True).count() - votes.filter(is_upvote=False).count()
    return render(request, 'dashboard/feedback/item_detail.html', {
        'item': item,
        'comments': comments,
        'votes': votes,
        'score': score,
        'status_choices': UnivFeedback.Status.choices,
    })


@login_required
@user_passes_test(is_admin)
def feedback_item_create(request):
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        category_id = request.POST.get('category_id') or None
        priority = request.POST.get('priority') or UnivFeedback.Priority.MEDIUM
        department_id = request.POST.get('department_id') or None
        course_id = request.POST.get('course_id') or None
        section_id = request.POST.get('section_id') or None
        faculty_id = request.POST.get('faculty_id') or None
        syllabus_id = request.POST.get('syllabus_id') or None
        if title and description:
            item = UnivFeedback.objects.create(
                title=title,
                description=description,
                category_id=category_id,
                created_by=request.user,
                priority=priority,
                department_id=department_id,
                course_id=course_id,
                section_id=section_id,
                faculty_id=faculty_id,
                syllabus_id=syllabus_id,
            )
            messages.success(request, 'Feedback created')
            return redirect('dashboard:feedback_item_detail', item_id=item.id)
        messages.error(request, 'Title and description are required')
    return render(request, 'dashboard/feedback/item_create.html', {
        'categories': UnivFeedbackCategory.objects.filter(is_active=True).order_by('name'),
        'departments': Department.objects.order_by('name'),
        'courses': Course.objects.order_by('code'),
        'sections': CourseSection.objects.order_by('course__code', 'section_number')[:200],
        'faculties': Faculty.objects.order_by('-created_at')[:200],
        'syllabi': Syllabus.objects.order_by('-created_at')[:200],
        'priority_choices': UnivFeedback.Priority.choices,
    })


@login_required
@user_passes_test(is_admin)
def dashboard_home(request):
    """Main dashboard view with statistics"""
    try:
        # Basic statistics that are most likely to work
        context = {
            'total_users': User.objects.count(),
            'total_students': Student.objects.count(),
            'total_faculty': Faculty.objects.count(),
            'total_roles': Role.objects.count(),
            'total_permissions': Permission.objects.count(),
            'active_sessions': UserSession.objects.filter(revoked=False).count(),
            'recent_users': User.objects.order_by('-date_joined')[:10],
        }
        
        # Try to add more statistics if possible
        try:
            context.update({
                'active_students': Student.objects.filter(status='ACTIVE').count(),
                'active_faculty': Faculty.objects.filter(status='ACTIVE', currently_associated=True).count(),
                'total_custom_fields': CustomField.objects.filter(is_active=True).count(),
                'total_faculty_custom_fields': FacultyCustomField.objects.filter(is_active=True).count(),
                'recent_students': Student.objects.order_by('-created_at')[:5],
                'recent_faculty': Faculty.objects.order_by('-created_at')[:5],
            })
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Some dashboard statistics could not be loaded: {str(e)}")
        
        return render(request, 'dashboard/home.html', context)
        
    except Exception as e:
        # If there's an error, return a simple dashboard with basic info
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in dashboard_home: {str(e)}")
        
        context = {
            'total_users': User.objects.count(),
            'total_students': Student.objects.count(),
            'total_faculty': Faculty.objects.count(),
            'error_message': 'Some statistics could not be loaded. Please check the logs.',
        }
        return render(request, 'dashboard/home.html', context)


@login_required
@user_passes_test(is_admin)
def facilities_dashboard(request):
    """Facilities management dashboard - redirects to facilities app"""
    return redirect('facilities_dashboard:dashboard')


@login_required
@user_passes_test(is_admin)
def users_list(request):
    """Users management page"""
    users = User.objects.all().order_by('-date_joined')
    return render(request, 'dashboard/users.html', {'users': users})


@login_required
@user_passes_test(is_admin)
def roles_list(request):
    """Roles management page"""
    roles = Role.objects.all().order_by('name')
    return render(request, 'dashboard/roles.html', {'roles': roles})


@login_required
@user_passes_test(is_admin)
def sessions_list(request):
    """Active sessions page"""
    sessions = UserSession.objects.all().order_by('-created_at')
    return render(request, 'dashboard/sessions.html', {'sessions': sessions})


@login_required
@user_passes_test(is_admin)
def audit_logs(request):
    """Audit logs page"""
    logs = AuditLog.objects.all().order_by('-created_at')[:100]
    return render(request, 'dashboard/audit_logs.html', {'logs': logs})


# ---------------------
# Grads & Marks Dashboard
# ---------------------

@login_required
@user_passes_test(is_admin)
def grads_dashboard(request):
    context = {
        'grade_scales_count': GradeScale.objects.count(),
        'terms_count': Term.objects.count(),
        'results_count': CourseResult.objects.count(),
        'students_with_records': GraduateRecord.objects.count(),
    }
    return render(request, 'dashboard/grads/dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def grads_grade_scales(request):
    items = GradeScale.objects.order_by('-grade_points')
    departments = Department.objects.order_by('code')
    programs = AcademicProgram.objects.order_by('code')
    return render(request, 'dashboard/grads/grade_scales.html', {
        'items': items,
        'departments': departments,
        'programs': programs,
    })


@login_required
@user_passes_test(is_admin)
def grads_terms(request):
    terms = Term.objects.order_by('-academic_year', 'semester')
    return render(request, 'dashboard/grads/terms.html', {'terms': terms})


# ---------------------
# Transportation Dashboard
# ---------------------

@login_required
@user_passes_test(is_admin)
def transport_dashboard(request):
    context = {
        'vehicles_count': Vehicle.objects.count(),
        'drivers_count': Driver.objects.count(),
        'routes_count': Route.objects.count(),
        'stops_count': Stop.objects.count(),
        'assignments_count': VehicleAssignment.objects.count(),
        'schedules_count': TripSchedule.objects.count(),
        'passes_count': TransportPass.objects.count(),
    }
    return render(request, 'dashboard/transportation/dashboard.html', context)


# -----------------
# Mentoring Dashboard
# -----------------

@login_required
@user_passes_test(is_admin)
def mentoring_dashboard(request):
    total_mentorships = Mentorship.objects.count()
    active_mentorships = Mentorship.objects.filter(is_active=True).count()
    total_projects = Project.objects.count()
    upcoming_meetings = Meeting.objects.filter(scheduled_at__gte=timezone.now()).order_by('scheduled_at')[:10]
    recent_feedback = Feedback.objects.order_by('-created_at')[:10]

    context = {
        'total_mentorships': total_mentorships,
        'active_mentorships': active_mentorships,
        'total_projects': total_projects,
        'upcoming_meetings': upcoming_meetings,
        'recent_feedback': recent_feedback,
    }
    return render(request, 'dashboard/mentoring/dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def mentoring_mentorships(request):
    mentorships = Mentorship.objects.select_related('mentor', 'student').all()[:200]
    return render(request, 'dashboard/mentoring/mentorships.html', {'mentorships': mentorships})


@login_required
@user_passes_test(is_admin)
def mentoring_projects(request):
    projects = Project.objects.select_related('mentorship').all()[:200]
    return render(request, 'dashboard/mentoring/projects.html', {'projects': projects})


@login_required
@user_passes_test(is_admin)
def mentoring_meetings(request):
    meetings = Meeting.objects.select_related('mentorship').order_by('-scheduled_at')[:200]
    return render(request, 'dashboard/mentoring/meetings.html', {'meetings': meetings})


@login_required
@user_passes_test(is_admin)
def mentoring_feedback(request):
    feedback = Feedback.objects.select_related('mentorship', 'project', 'meeting').order_by('-created_at')[:200]
    return render(request, 'dashboard/mentoring/feedback.html', {'feedback': feedback})


@login_required
@user_passes_test(is_admin)
def transport_vehicles(request):
    q = request.GET.get('q')
    items = Vehicle.objects.all()
    if q:
        items = items.filter(Q(number_plate__icontains=q) | Q(registration_number__icontains=q) | Q(make__icontains=q) | Q(model__icontains=q))
    items = items.order_by('number_plate')
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="vehicles.csv"'
        writer = csv.writer(response)
        writer.writerow(['Number Plate', 'Registration', 'Make', 'Model', 'Capacity', 'Active'])
        for v in items:
            writer.writerow([v.number_plate, v.registration_number, v.make, v.model, v.capacity, v.is_active])
        return response
    return render(request, 'dashboard/transportation/vehicles.html', {'items': items, 'q': q or ''})


@login_required
@user_passes_test(is_admin)
def transport_drivers(request):
    q = request.GET.get('q')
    items = Driver.objects.all()
    if q:
        items = items.filter(Q(full_name__icontains=q) | Q(phone__icontains=q) | Q(license_number__icontains=q))
    items = items.order_by('full_name')
    return render(request, 'dashboard/transportation/drivers.html', {'items': items, 'q': q or ''})


@login_required
@user_passes_test(is_admin)
def transport_routes(request):
    q = request.GET.get('q')
    items = Route.objects.all()
    if q:
        items = items.filter(Q(name__icontains=q) | Q(description__icontains=q))
    items = items.order_by('name').prefetch_related('route_stops__stop')
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="routes.csv"'
        writer = csv.writer(response)
        writer.writerow(['Name', 'Description', 'Distance (km)', 'Active'])
        for r in items:
            writer.writerow([r.name, r.description, r.distance_km, r.is_active])
        return response
    return render(request, 'dashboard/transportation/routes.html', {'items': items, 'q': q or ''})


@login_required
@user_passes_test(is_admin)
def transport_stops(request):
    q = request.GET.get('q')
    items = Stop.objects.all()
    if q:
        items = items.filter(Q(name__icontains=q) | Q(landmark__icontains=q))
    items = items.order_by('name')
    return render(request, 'dashboard/transportation/stops.html', {'items': items, 'q': q or ''})


@login_required
@user_passes_test(is_admin)
def transport_assignments(request):
    q = request.GET.get('q')
    items = VehicleAssignment.objects.select_related('vehicle', 'driver', 'route').all()
    if q:
        items = items.filter(Q(vehicle__number_plate__icontains=q) | Q(driver__full_name__icontains=q) | Q(route__name__icontains=q))
    items = items.order_by('-start_date')
    return render(request, 'dashboard/transportation/assignments.html', {'items': items, 'q': q or ''})


@login_required
@user_passes_test(is_admin)
def transport_schedules(request):
    q = request.GET.get('q')
    items = TripSchedule.objects.select_related('assignment', 'assignment__vehicle', 'assignment__route').all()
    if q:
        items = items.filter(Q(assignment__vehicle__number_plate__icontains=q) | Q(assignment__route__name__icontains=q))
    items = items.order_by('day_of_week', 'departure_time')
    return render(request, 'dashboard/transportation/schedules.html', {'items': items, 'q': q or ''})


@login_required
@user_passes_test(is_admin)
def transport_passes(request):
    q = request.GET.get('q')
    items = TransportPass.objects.select_related('user', 'route', 'start_stop', 'end_stop').all()
    if q:
        items = items.filter(Q(user__username__icontains=q) | Q(user__email__icontains=q) | Q(route__name__icontains=q))
    items = items.order_by('-valid_from')
    return render(request, 'dashboard/transportation/passes.html', {'items': items, 'q': q or ''})


# Create Views (simple create-only forms for now)

@login_required
@user_passes_test(is_admin)
def transport_vehicle_create(request):
    if request.method == 'POST':
        form = VehicleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vehicle created successfully.')
            return redirect('dashboard:transport_vehicles')
    else:
        form = VehicleForm()
    return render(request, 'dashboard/transportation/create_form.html', {'form': form, 'title': 'Add Vehicle'})


@login_required
@user_passes_test(is_admin)
def transport_driver_create(request):
    if request.method == 'POST':
        form = DriverForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Driver created successfully.')
            return redirect('dashboard:transport_drivers')
    else:
        form = DriverForm()
    return render(request, 'dashboard/transportation/create_form.html', {'form': form, 'title': 'Add Driver'})


@login_required
@user_passes_test(is_admin)
def transport_route_create(request):
    if request.method == 'POST':
        form = RouteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Route created successfully.')
            return redirect('dashboard:transport_routes')
    else:
        form = RouteForm()
    return render(request, 'dashboard/transportation/create_form.html', {'form': form, 'title': 'Add Route'})


@login_required
@user_passes_test(is_admin)
def transport_stop_create(request):
    if request.method == 'POST':
        form = StopForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Stop created successfully.')
            return redirect('dashboard:transport_stops')
    else:
        form = StopForm()
    return render(request, 'dashboard/transportation/create_form.html', {'form': form, 'title': 'Add Stop'})


@login_required
@user_passes_test(is_admin)
def transport_assignment_create(request):
    if request.method == 'POST':
        form = VehicleAssignmentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Assignment created successfully.')
            return redirect('dashboard:transport_assignments')
    else:
        form = VehicleAssignmentForm()
    return render(request, 'dashboard/transportation/create_form.html', {'form': form, 'title': 'Add Assignment'})


@login_required
@user_passes_test(is_admin)
def transport_schedule_create(request):
    if request.method == 'POST':
        form = TripScheduleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Schedule created successfully.')
            return redirect('dashboard:transport_schedules')
    else:
        form = TripScheduleForm()
    return render(request, 'dashboard/transportation/create_form.html', {'form': form, 'title': 'Add Schedule'})


@login_required
@user_passes_test(is_admin)
def transport_pass_create(request):
    if request.method == 'POST':
        form = TransportPassForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Transport pass created successfully.')
            return redirect('dashboard:transport_passes')
    else:
        form = TransportPassForm()
    return render(request, 'dashboard/transportation/create_form.html', {'form': form, 'title': 'Add Transport Pass'})


@login_required
@user_passes_test(is_admin)
def grads_results(request):
    student_id = request.GET.get('student')
    term_id = request.GET.get('term')
    dept_code = request.GET.get('department')
    course_id = request.GET.get('course')
    year = request.GET.get('year')
    section = request.GET.get('section')
    results = CourseResult.objects.select_related('student', 'term', 'course_section', 'course_section__course')
    if student_id:
        results = results.filter(student_id=student_id)
    if term_id:
        results = results.filter(term_id=term_id)
    if dept_code:
        results = results.filter(course_section__course__department__code__iexact=dept_code)
    if course_id:
        results = results.filter(course_section__course_id=course_id)
    if year:
        results = results.filter(course_section__academic_year=year)
    if section:
        results = results.filter(course_section__section_number__iexact=section)
    results = results.order_by('-evaluated_at')[:500]
    # Build filter option lists
    years = CourseSection.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')
    sections = CourseSection.objects.values_list('section_number', flat=True).distinct().order_by('section_number')
    courses = Course.objects.order_by('code')[:500]
    students_qs = Student.objects.order_by('roll_number')[:500]
    course_sections = CourseSection.objects.select_related('course').order_by('course__code', 'section_number')[:500]
    return render(request, 'dashboard/grads/results.html', {
        'results': results,
        'terms': Term.objects.all().order_by('-academic_year'),
        'years': years,
        'sections': sections,
        'courses': courses,
        'students_list': students_qs,
        'course_sections': course_sections,
    })


@login_required
@user_passes_test(is_admin)
def grads_bulk_entry(request):
    terms = Term.objects.all().order_by('-academic_year')
    years = CourseSection.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')
    sections = CourseSection.objects.values_list('section_number', flat=True).distinct().order_by('section_number')
    courses = Course.objects.order_by('code')[:500]
    students_qs = Student.objects.order_by('roll_number')[:500]
    return render(request, 'dashboard/grads/bulk_entry.html', {
        'terms': terms,
        'years': years,
        'sections': sections,
        'courses': courses,
        'students_list': students_qs,
    })


@login_required
@user_passes_test(is_admin)
def grads_sections_api(request):
    year = request.GET.get('year')
    course_id = request.GET.get('course')
    qs = CourseSection.objects.all()
    if year:
        qs = qs.filter(academic_year=year)
    if course_id:
        qs = qs.filter(course_id=course_id)
    data = [
        {
            'id': s.id,
            'section_number': s.section_number,
            'faculty': getattr(s.faculty, 'name', ''),
            'semester': s.semester,
        } for s in qs.order_by('section_number')[:200]
    ]
    return JsonResponse({'sections': data})


@login_required
@user_passes_test(is_admin)
def grads_students_api(request):
    year = request.GET.get('year')
    section = request.GET.get('section')
    # Filter students by enrollment history if available else return all
    students = Student.objects.all()
    if section:
        students = students.filter(section=section)
    if year:
        students = students.filter(academic_year=year)
    data = [
        {
            'id': s.id,
            'roll_number': s.roll_number,
            'full_name': s.full_name,
        } for s in students.order_by('roll_number')[:500]
    ]
    return JsonResponse({'students': data})


@login_required
@user_passes_test(is_admin)
def grads_transcript(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    grad = GraduateRecord.objects.filter(student=student).first()
    term_gpas = TermGPA.objects.filter(student=student).select_related('term').order_by('-term__academic_year')
    results = CourseResult.objects.filter(student=student).select_related('term', 'course_section', 'course_section__course').order_by('-term__academic_year')
    context = {
        'student': student,
        'grad': grad,
        'term_gpas': term_gpas,
        'results': results,
    }
    return render(request, 'dashboard/grads/transcript.html', context)


@login_required
@user_passes_test(is_admin)
def database_schema(request):
    """Database schema overview"""
    with connection.cursor() as cursor:
        # PostgreSQL-compatible query to get table names
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_type = 'BASE TABLE'
            ORDER BY table_name;
        """)
        tables = [row[0] for row in cursor.fetchall()]
    
    schema_info = []
    for table in tables:
        with connection.cursor() as cursor:
            # PostgreSQL-compatible query to get column information
            cursor.execute("""
                SELECT 
                    column_name,
                    data_type,
                    is_nullable,
                    column_default,
                    character_maximum_length
                FROM information_schema.columns 
                WHERE table_name = %s 
                AND table_schema = 'public'
                ORDER BY ordinal_position;
            """, [table])
            columns = cursor.fetchall()
            schema_info.append({
                'table': table,
                'columns': columns
            })
    
    return render(request, 'dashboard/schema.html', {'schema_info': schema_info})


# API Endpoints for dashboard data
@api_view(['GET'])
@permission_classes([IsAdminUser])
def api_database_schema(request):
    """API endpoint to get database schema information"""
    with connection.cursor() as cursor:
        # PostgreSQL-compatible query to get table names
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_type = 'BASE TABLE'
            ORDER BY table_name;
        """)
        tables = [row[0] for row in cursor.fetchall()]
    
    schema_data = {}
    for table in tables:
        with connection.cursor() as cursor:
            # PostgreSQL-compatible query to get column information
            cursor.execute("""
                SELECT 
                    column_name,
                    data_type,
                    is_nullable,
                    column_default,
                    character_maximum_length
                FROM information_schema.columns 
                WHERE table_name = %s 
                AND table_schema = 'public'
                ORDER BY ordinal_position;
            """, [table])
            columns = cursor.fetchall()
            schema_data[table] = [
                {
                    'name': col[0],
                    'type': col[1],
                    'not_null': col[2] == 'NO',
                    'default': col[3],
                    'max_length': col[4]
                }
                for col in columns
            ]
    
    return Response(schema_data)


@login_required
@user_passes_test(is_admin)
def download_schema_excel(request):
    """Download database schema as an Excel file with one sheet per table."""
    # Lazy import to avoid hard dependency if not used
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse("openpyxl is not installed. Please install it to use Excel export.", status=500)
    except Exception as exc:
        return HttpResponse(f"Error importing openpyxl: {exc}", status=500)

    wb = Workbook()
    wb.remove(wb.active)

    with connection.cursor() as cursor:
        # PostgreSQL-compatible query to get table names
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_type = 'BASE TABLE'
            ORDER BY table_name;
        """)
        tables = [row[0] for row in cursor.fetchall()]

    for table in tables:
        ws = wb.create_sheet(title=str(table)[:31])  # Excel sheet name limit
        ws.append(["column", "type", "not_null", "default", "max_length"])  
        with connection.cursor() as cursor:
            # PostgreSQL-compatible query to get column information
            cursor.execute("""
                SELECT 
                    column_name,
                    data_type,
                    is_nullable,
                    column_default,
                    character_maximum_length
                FROM information_schema.columns 
                WHERE table_name = %s 
                AND table_schema = 'public'
                ORDER BY ordinal_position;
            """, [table])
            for col in cursor.fetchall():
                ws.append([col[0], col[1], col[2] == 'NO', col[3], col[4]])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="database_schema.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def download_schema_excel_single(request):
    """Download database schema as a single-sheet Excel file consolidating all tables."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse("openpyxl is not installed. Please install it to use Excel export.", status=500)
    except Exception as exc:
        return HttpResponse(f"Error importing openpyxl: {exc}", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = 'schema'
    ws.append(["table", "column", "type", "not_null", "default", "max_length"])  

    with connection.cursor() as cursor:
        # PostgreSQL-compatible query to get table names
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_type = 'BASE TABLE'
            ORDER BY table_name;
        """)
        tables = [row[0] for row in cursor.fetchall()]

    for table in tables:
        with connection.cursor() as cursor:
            # PostgreSQL-compatible query to get column information
            cursor.execute("""
                SELECT 
                    column_name,
                    data_type,
                    is_nullable,
                    column_default,
                    character_maximum_length
                FROM information_schema.columns 
                WHERE table_name = %s 
                AND table_schema = 'public'
                ORDER BY ordinal_position;
            """, [table])
            for col in cursor.fetchall():
                ws.append([table, col[0], col[1], col[2] == 'NO', col[3], col[4]])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="database_schema_single.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def download_schema_csv(request):
    """Download database schema as CSV with all tables consolidated."""
    try:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="database_schema.csv"'

        writer = csv.writer(response)
        writer.writerow(["table", "column", "type", "not_null", "default", "max_length"])  

        with connection.cursor() as cursor:
            # PostgreSQL-compatible query to get table names
            cursor.execute("""
                SELECT table_name 
                FROM information_schema.tables 
                WHERE table_schema = 'public' 
                AND table_type = 'BASE TABLE'
                ORDER BY table_name;
            """)
            tables = [row[0] for row in cursor.fetchall()]

        for table in tables:
            with connection.cursor() as cursor:
                # PostgreSQL-compatible query to get column information
                cursor.execute("""
                    SELECT 
                        column_name,
                        data_type,
                        is_nullable,
                        column_default,
                        character_maximum_length
                    FROM information_schema.columns 
                    WHERE table_name = %s 
                    AND table_schema = 'public'
                    ORDER BY ordinal_position;
                """, [table])
                for col in cursor.fetchall():
                    writer.writerow([table, col[0], col[1], col[2] == 'NO', col[3], col[4]])

        return response
    except Exception as e:
        return HttpResponse(f"Error generating CSV: {e}", status=500)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def api_table_data(request, table_name):
    """API endpoint to get data from any table"""
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT * FROM {table_name} LIMIT 100;")
            columns = [desc[0] for desc in cursor.description]
            rows = cursor.fetchall()
            
            data = []
            for row in rows:
                data.append(dict(zip(columns, row)))
        
        return Response({
            'table': table_name,
            'columns': columns,
            'data': data,
            'count': len(data)
        })
    except Exception as e:
        return Response({'error': str(e)}, status=400)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def api_dashboard_stats(request):
    """API endpoint for dashboard statistics"""
    stats = {
        'users': {
            'total': User.objects.count(),
            'active': User.objects.filter(is_active=True).count(),
            'staff': User.objects.filter(is_staff=True).count(),
            'verified': User.objects.filter(is_verified=True).count(),
        },
        'auth': {
            'roles': Role.objects.count(),
            'permissions': Permission.objects.count(),
            'identifiers': AuthIdentifier.objects.count(),
            'active_sessions': UserSession.objects.filter(revoked=False).count(),
        },
        'security': {
            'failed_logins': FailedLogin.objects.count(),
            'audit_logs': AuditLog.objects.count(),
        }
    }
    return Response(stats)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def api_models_info(request):
    """API endpoint to get information about all Django models"""
    models_info = {}
    
    for model in apps.get_models():
        app_label = model._meta.app_label
        model_name = model._meta.model_name
        
        if app_label not in models_info:
            models_info[app_label] = {}
        
        fields_info = []
        for field in model._meta.fields:
            fields_info.append({
                'name': field.name,
                'type': field.__class__.__name__,
                'null': field.null,
                'blank': field.blank,
                'unique': field.unique,
            })
        
        models_info[app_label][model_name] = {
            'table_name': model._meta.db_table,
            'fields': fields_info,
            'count': model.objects.count() if hasattr(model.objects, 'count') else 0,
        }
    
    return Response(models_info)


# ------------------------------
# ER diagram generation
# ------------------------------
def _generate_mermaid_er_diagram() -> str:
    """Build Mermaid ER diagram from Django model metadata."""
    # Header
    lines = ["erDiagram"]

    # Collect nodes
    all_models = list(apps.get_models())
    for model in all_models:
        model_label = model.__name__
        lines.append(f"    {model_label} {{")
        for field in model._meta.get_fields():
            # Skip reverse relations rendered later via FK fields
            if hasattr(field, 'remote_field') and field.auto_created and not field.concrete:
                continue
            if getattr(field, 'many_to_many', False):
                continue
            # Field name and type
            field_name = getattr(field, 'name', 'id')
            field_type = getattr(field, 'get_internal_type', lambda: field.__class__.__name__)()
            # Mark PK
            pk_suffix = " PK" if getattr(field, 'primary_key', False) else ""
            lines.append(f"        {field_type} {field_name}{pk_suffix}")
        lines.append("    }")

    # Collect relations from FK and O2O
    for model in all_models:
        for field in model._meta.get_fields():
            if getattr(field, 'many_to_many', False):
                # Many-to-many: A }o--o{ B
                try:
                    target = field.related_model
                    if target is None:
                        continue
                    a = model.__name__
                    b = target.__name__
                    lines.append(f"    {a} }}o--o{{ {b} : many_to_many")
                except Exception:
                    continue
            elif getattr(field, 'many_to_one', False) or getattr(field, 'one_to_one', False):
                # For FK (many-to-one): Parent ||--o{ Child
                try:
                    target = field.related_model
                    if target is None:
                        continue
                    parent = target.__name__
                    child = model.__name__
                    connector = "||--||" if getattr(field, 'one_to_one', False) else "||--o{"
                    rel_name = getattr(field, 'name', 'fk')
                    lines.append(f"    {parent} {connector} {child} : {rel_name}")
                except Exception:
                    continue

    return "\n".join(lines)


@login_required
@user_passes_test(is_admin)
def er_diagram_page(request):
    """Render a page that shows the ER diagram using Mermaid."""
    mermaid = _generate_mermaid_er_diagram()
    return render(request, 'dashboard/er.html', { 'mermaid': mermaid })


# ------------------------------
# R&D (Research & Development)
# ------------------------------

@login_required
@user_passes_test(is_admin)
def rnd_dashboard(request):
    context = {
        'projects_count': RndProject.objects.count(),
        'publications_count': RndPublication.objects.count(),
        'grants_count': RndGrant.objects.count(),
        'researchers_count': RndResearcher.objects.count(),
        'datasets_count': RndDataset.objects.count(),
        'patents_count': RndPatent.objects.count(),
        'recent_projects': RndProject.objects.order_by('-start_date')[:10],
        'recent_publications': RndPublication.objects.order_by('-year')[:10],
    }
    return render(request, 'dashboard/rnd/dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def rnd_projects(request):
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    qs = RndProject.objects.select_related('principal_investigator').prefetch_related('members')
    if search:
        qs = qs.filter(Q(title__icontains=search) | Q(abstract__icontains=search))
    if status:
        qs = qs.filter(status=status)
    projects = qs.order_by('-start_date')[:500]
    return render(request, 'dashboard/rnd/projects.html', {
        'projects': projects,
        'status_filter': status,
        'search': search,
    })


@login_required
@user_passes_test(is_admin)
def rnd_project_detail(request, project_id: int):
    project = get_object_or_404(RndProject.objects.select_related('principal_investigator').prefetch_related('members', 'grants', 'publications', 'datasets', 'patents', 'collaborations'), id=project_id)
    return render(request, 'dashboard/rnd/project_detail.html', {
        'project': project,
    })


@login_required
@user_passes_test(is_admin)
def rnd_researchers(request):
    search = request.GET.get('search', '')
    dept = request.GET.get('department', '')
    qs = RndResearcher.objects.select_related('user')
    if search:
        qs = qs.filter(Q(user__first_name__icontains=search) | Q(user__last_name__icontains=search) | Q(user__username__icontains=search) | Q(department__icontains=search))
    if dept:
        qs = qs.filter(department__icontains=dept)
    researchers = qs.order_by('user__last_name', 'user__first_name')[:500]
    return render(request, 'dashboard/rnd/researchers.html', {
        'researchers': researchers,
        'search': search,
        'department': dept,
    })


@login_required
@user_passes_test(is_admin)
def rnd_researcher_detail(request, researcher_id: int):
    researcher = get_object_or_404(RndResearcher.objects.select_related('user'), id=researcher_id)
    led_projects = RndProject.objects.filter(principal_investigator=researcher).order_by('-start_date')
    member_projects = RndProject.objects.filter(members=researcher).exclude(principal_investigator=researcher).order_by('-start_date')
    publications = RndPublication.objects.filter(authors=researcher).order_by('-year')
    patents = RndPatent.objects.filter(inventors=researcher).order_by('-grant_date')
    return render(request, 'dashboard/rnd/researcher_detail.html', {
        'researcher': researcher,
        'led_projects': led_projects,
        'member_projects': member_projects,
        'publications': publications,
        'patents': patents,
    })


@login_required
@user_passes_test(is_admin)
def rnd_grants(request):
    grants = RndGrant.objects.order_by('-start_date')[:500]
    return render(request, 'dashboard/rnd/grants.html', {'grants': grants})


@login_required
@user_passes_test(is_admin)
def rnd_publications(request):
    pubs = RndPublication.objects.order_by('-year', 'title')[:500]
    return render(request, 'dashboard/rnd/publications.html', {'publications': pubs})


@login_required
@user_passes_test(is_admin)
def rnd_patents(request):
    items = RndPatent.objects.order_by('-grant_date', '-filing_date')[:500]
    return render(request, 'dashboard/rnd/patents.html', {'patents': items})


@login_required
@user_passes_test(is_admin)
def rnd_datasets(request):
    items = RndDataset.objects.order_by('name')[:500]
    return render(request, 'dashboard/rnd/datasets.html', {'datasets': items})


@login_required
@user_passes_test(is_admin)
def rnd_collaborations(request):
    items = RndCollaboration.objects.select_related('project').order_by('-start_date')[:500]
    return render(request, 'dashboard/rnd/collaborations.html', {'collaborations': items})


@api_view(['GET'])
@permission_classes([IsAdminUser])
def api_er_diagram(request):
    """Return Mermaid ER diagram string."""
    mermaid = _generate_mermaid_er_diagram()
    return Response({ 'mermaid': mermaid })


@login_required
@user_passes_test(is_admin)
def test_openpyxl(request):
    """Test endpoint to check openpyxl import."""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        return HttpResponse("openpyxl works fine!")
    except ImportError as e:
        return HttpResponse(f"ImportError: {e}", status=500)
    except Exception as e:
        return HttpResponse(f"Other error: {e}", status=500)


# ------------------------------
# Student Management Views
# ------------------------------
@login_required
@user_passes_test(is_admin)
def students_list(request):
    """Students management page"""
    from students.models import Student
    
    # Get filter parameters
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    grade = request.GET.get('grade', '')
    section = request.GET.get('section', '')
    
    # Build queryset
    students = Student.objects.all()
    
    if search:
        students = students.filter(
            Q(roll_number__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search) |
            Q(father_name__icontains=search) |
            Q(mother_name__icontains=search)
        )
    
    if status:
        students = students.filter(status=status)
    
    if grade:
        students = students.filter(grade_level=grade)
    
    if section:
        students = students.filter(section=section)
    
    students = students.order_by('-created_at')
    
    # Get choices for filter
    from students.models import Student
    year_choices = Student.YEAR_OF_STUDY_CHOICES
    semester_choices = Student.SEMESTER_CHOICES
    
    context = {
        'students': students,
        'year_choices': year_choices,
        'semester_choices': semester_choices,
    }
    return render(request, 'dashboard/students.html', context)


@login_required
@user_passes_test(is_admin)
def student_detail(request, student_id):
    """Student detail page"""
    try:
        student = Student.objects.get(id=student_id)
        context = {
            'student': student,
            'enrollment_history': student.enrollment_history.all(),
            'documents': student.documents.all(),
            'custom_fields': student.custom_field_values.all(),
        }
        return render(request, 'dashboard/student_detail.html', context)
    except Student.DoesNotExist:
        return render(request, 'dashboard/404.html', status=404)


@login_required
@user_passes_test(is_admin)
def custom_fields_list(request):
    """Custom fields management page"""
    from students.models import CustomField
    from django.utils import timezone
    
    custom_fields = CustomField.objects.all().order_by('order', 'name')
    
    # Calculate statistics
    active_fields_count = custom_fields.filter(is_active=True).count()
    required_fields_count = custom_fields.filter(required=True).count()
    field_types_count = custom_fields.values('field_type').distinct().count()
    
    context = {
        'custom_fields': custom_fields,
        'active_fields_count': active_fields_count,
        'required_fields_count': required_fields_count,
        'field_types_count': field_types_count,
        'now': timezone.now(),
    }
    return render(request, 'dashboard/custom_fields.html', context)


@login_required
@user_passes_test(is_admin)
def student_login_page(request):
    """Student login page"""
    return render(request, 'dashboard/student_login.html')


@login_required
@user_passes_test(is_admin)
def student_sessions(request):
    """Student login sessions page"""
    from django.utils import timezone
    from datetime import timedelta
    
    # Get filter parameters
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Build queryset
    sessions = UserSession.objects.filter(user__student_profile__isnull=False)
    
    if search:
        sessions = sessions.filter(
            Q(user__student_profile__roll_number__icontains=search) |
            Q(user__student_profile__first_name__icontains=search) |
            Q(user__student_profile__last_name__icontains=search) |
            Q(user__email__icontains=search)
        )
    
    if status == 'active':
        sessions = sessions.filter(revoked=False, expires_at__gt=timezone.now())
    elif status == 'expired':
        sessions = sessions.filter(expires_at__lt=timezone.now())
    elif status == 'revoked':
        sessions = sessions.filter(revoked=True)
    
    if date_from:
        sessions = sessions.filter(created_at__date__gte=date_from)
    
    if date_to:
        sessions = sessions.filter(created_at__date__lte=date_to)
    
    sessions = sessions.order_by('-created_at')
    
    # Calculate statistics
    active_sessions_count = UserSession.objects.filter(
        user__student_profile__isnull=False,
        revoked=False,
        expires_at__gt=timezone.now()
    ).count()
    
    unique_students_count = UserSession.objects.filter(
        user__student_profile__isnull=False
    ).values('user').distinct().count()
    
    today = timezone.now().date()
    today_logins_count = UserSession.objects.filter(
        user__student_profile__isnull=False,
        created_at__date=today
    ).count()
    
    context = {
        'sessions': sessions,
        'active_sessions_count': active_sessions_count,
        'unique_students_count': unique_students_count,
        'today_logins_count': today_logins_count,
        'now': timezone.now(),
    }
    return render(request, 'dashboard/student_sessions.html', context)


# ------------------------------
# Student Import Views
# ------------------------------
@login_required
@user_passes_test(is_admin)
def student_import_page(request):
    """Student import page"""
    recent_imports = StudentImport.objects.all()[:10]
    context = {
        'recent_imports': recent_imports,
    }
    return render(request, 'dashboard/student_import.html', context)


@login_required
@user_passes_test(is_admin)
def student_import_process(request):
    """Process student import from Excel/CSV file"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'success': False, 'error': 'No file uploaded'})
        
        # Validate file type
        allowed_extensions = ['.xlsx', '.xls', '.csv']
        file_extension = os.path.splitext(file.name)[1].lower()
        if file_extension not in allowed_extensions:
            return JsonResponse({'success': False, 'error': 'Invalid file type. Please upload Excel or CSV file.'})
        
        # Get import options
        skip_errors = request.POST.get('skip_errors') == 'on'
        create_login = request.POST.get('create_login') == 'on'
        update_existing = request.POST.get('update_existing') == 'on'
        
        # Create import record
        import_record = StudentImport.objects.create(
            filename=file.name,
            file_size=file.size,
            created_by=request.user,
            skip_errors=skip_errors,
            create_login=create_login,
            update_existing=update_existing,
            status='PROCESSING'
        )
        
        # Process the file
        result = process_student_import(file, import_record, skip_errors, create_login, update_existing)
        
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@user_passes_test(is_admin)
def download_template(request):
    """Download Excel template for student import"""
    try:
        import pandas as pd
        from io import BytesIO
        
        # Create sample data
        sample_data = {
            'roll_number': ['STU001', 'STU002', 'STU003'],
            'first_name': ['John', 'Jane', 'Mike'],
            'last_name': ['Doe', 'Smith', 'Johnson'],
            'middle_name': ['A', 'B', 'C'],
            'date_of_birth': ['2010-01-15', '2010-03-20', '2010-06-10'],
            'gender': ['M', 'F', 'M'],
            'section': ['A', 'B', 'A'],
            'academic_year': ['2023-2024', '2023-2024', '2023-2024'],
            'email': ['john.doe@example.com', 'jane.smith@example.com', 'mike.johnson@example.com'],
            'student_mobile': ['+1234567890', '+1234567891', '+1234567892'],
            'quota': ['GENERAL', 'SC', 'OBC'],
            'rank': ['1', '5', '10'],
            'status': ['ACTIVE', 'ACTIVE', 'ACTIVE'],
            'father_name': ['John Doe Sr', 'James Smith', 'Robert Johnson'],
            'mother_name': ['Mary Doe', 'Sarah Smith', 'Lisa Johnson'],
            'father_mobile': ['+1234567893', '+1234567894', '+1234567895'],
            'mother_mobile': ['+1234567896', '+1234567897', '+1234567898'],
            'address': ['123 Main St', '456 Oak Ave', '789 Pine Rd'],
            'city': ['New York', 'Los Angeles', 'Chicago'],
            'state': ['NY', 'CA', 'IL'],
            'country': ['USA', 'USA', 'USA'],
            'postal_code': ['10001', '90210', '60601'],
        }
        
        # Create DataFrame
        df = pd.DataFrame(sample_data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Students', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Students']
            
            # Add instructions sheet
            instructions_data = {
                'Field': [
                    'roll_number', 'first_name', 'last_name', 'middle_name', 'date_of_birth', 'gender',
                    'section', 'academic_year', 'email', 'student_mobile', 'quota', 'rank',
                    'status', 'father_name', 'mother_name', 'father_mobile', 'mother_mobile', 'address',
                    'city', 'state', 'country', 'postal_code'
                ],
                'Required': [
                    'Yes', 'Yes', 'Yes', 'No', 'Yes', 'Yes', 'No', 'No', 'No', 'No', 'No', 'No', 'No',
                    'No', 'No', 'No', 'No', 'No', 'No', 'No', 'No', 'No', 'No'
                ],
                'Description': [
                    'Unique student identifier', 'Student first name', 'Student last name', 'Student middle name',
                    'Date of birth (YYYY-MM-DD)', 'Gender: M (Male), F (Female), O (Other)',
                    'Section: A, B, C, D, E', 'Academic year (e.g., 2023-2024)',
                    'Student email address', 'Student phone number', 'Quota category', 'Academic rank',
                    'Status: ACTIVE, INACTIVE, GRADUATED', 'Father name', 'Mother name', 'Father phone',
                    'Mother phone', 'Full address', 'City', 'State', 'Country', 'Postal code'
                ],
                'Example': [
                    'STU001', 'John', 'Doe', 'A', '2010-01-15', 'M', 'A', '2023-2024',
                    'john.doe@example.com', '+1234567890', 'GENERAL', '1', 'ACTIVE', 'John Doe Sr',
                    'Mary Doe', '+1234567893', '+1234567896', '123 Main St', 'New York', 'NY', 'USA', '10001'
                ]
            }
            
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
        
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
        return response
        
    except ImportError:
        return JsonResponse({'error': 'openpyxl is not installed'}, status=500)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def process_student_import(file, import_record, skip_errors, create_login, update_existing):
    """Process student import from file"""
    try:
        import pandas as pd
        from datetime import datetime
        
        # Read file based on extension
        file_extension = os.path.splitext(file.name)[1].lower()
        
        if file_extension == '.csv':
            df = pd.read_csv(file)
        else:  # Excel files
            df = pd.read_excel(file)
        
        # Remove empty rows
        df = df.dropna(how='all')
        
        # Initialize counters
        success_count = 0
        error_count = 0
        warning_count = 0
        errors = []
        warnings = []
        
        # Update import record
        import_record.total_rows = len(df)
        import_record.save()
        
        # Process each row
        for index, row in df.iterrows():
            row_number = index + 2  # +2 because Excel is 1-indexed and we have header
            
            try:
                # Validate required fields
                required_fields = ['roll_number', 'first_name', 'last_name', 'date_of_birth', 'gender']
                for field in required_fields:
                    if pd.isna(row.get(field, '')) or str(row[field]).strip() == '':
                        raise ValueError(f"Required field '{field}' is missing or empty")
                
                # Clean and validate data
                roll_number = str(row['roll_number']).strip()
                first_name = str(row['first_name']).strip()
                last_name = str(row['last_name']).strip()
                date_of_birth = str(row['date_of_birth']).strip()
                gender = str(row['gender']).strip().upper()
                
                # Validate gender
                if gender not in ['M', 'F', 'O']:
                    raise ValueError(f"Invalid gender '{gender}'. Must be M, F, or O")
                
                # Validate date format
                try:
                    datetime.strptime(date_of_birth, '%Y-%m-%d')
                except ValueError:
                    raise ValueError(f"Invalid date format '{date_of_birth}'. Use YYYY-MM-DD")
                
                # Check if student exists
                existing_student = None
                if update_existing:
                    existing_student = Student.objects.filter(roll_number=roll_number).first()
                
                if existing_student:
                    # Update existing student
                    student = existing_student
                    warnings.append({
                        'row': row_number,
                        'field': 'roll_number',
                        'message': f"Student with roll number '{roll_number}' already exists. Updating."
                    })
                    warning_count += 1
                else:
                    # Create new student
                    student = Student()
                
                # Set basic fields
                student.roll_number = roll_number
                student.first_name = first_name
                student.last_name = last_name
                student.date_of_birth = date_of_birth
                student.gender = gender
                # grade_level removed in schema; ignore if present in file
                
                # Set optional fields
                if not pd.isna(row.get('middle_name', '')):
                    student.middle_name = str(row['middle_name']).strip()
                
                if not pd.isna(row.get('email', '')):
                    student.email = str(row['email']).strip()
                
                if not pd.isna(row.get('student_mobile', '')):
                    student.student_mobile = str(row['student_mobile']).strip()
                
                if not pd.isna(row.get('section', '')):
                    section = str(row['section']).strip().upper()
                    if section in ['A', 'B', 'C', 'D', 'E']:
                        student.section = section
                
                if not pd.isna(row.get('academic_year', '')):
                    student.academic_year = str(row['academic_year']).strip()
                
                if not pd.isna(row.get('quota', '')):
                    quota = str(row['quota']).strip().upper()
                    if quota in [choice[0] for choice in Student.QUOTA_CHOICES]:
                        student.quota = quota
                
                if not pd.isna(row.get('rank', '')):
                    try:
                        student.rank = int(row['rank'])
                    except (ValueError, TypeError):
                        pass
                
                if not pd.isna(row.get('status', '')):
                    status = str(row['status']).strip().upper()
                    if status in [choice[0] for choice in Student.STATUS_CHOICES]:
                        student.status = status
                
                # Set parent information
                if not pd.isna(row.get('father_name', '')):
                    student.father_name = str(row['father_name']).strip()
                
                if not pd.isna(row.get('mother_name', '')):
                    student.mother_name = str(row['mother_name']).strip()
                
                if not pd.isna(row.get('father_mobile', '')):
                    student.father_mobile = str(row['father_mobile']).strip()
                
                if not pd.isna(row.get('mother_mobile', '')):
                    student.mother_mobile = str(row['mother_mobile']).strip()
                
                # Set address information
                address_parts = []
                for field in ['address', 'city', 'state', 'country', 'postal_code']:
                    if not pd.isna(row.get(field, '')):
                        address_parts.append(str(row[field]).strip())
                
                if address_parts:
                    student.full_address = ', '.join(address_parts)
                
                # Save student
                student.save()

                # Ensure login/identifiers if requested
                if create_login:
                    try:
                        if student.user:
                            from accounts.models import AuthIdentifier, IdentifierType
                            if student.email:
                                AuthIdentifier.objects.get_or_create(
                                    user=student.user,
                                    identifier=student.email,
                                    id_type=IdentifierType.EMAIL,
                                    defaults={'is_primary': True}
                                )
                            AuthIdentifier.objects.get_or_create(
                                user=student.user,
                                identifier=student.roll_number,
                                id_type=IdentifierType.USERNAME,
                                defaults={'is_primary': not bool(student.email)}
                            )
                    except Exception:
                        pass

                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append({
                    'row': row_number,
                    'field': 'general',
                    'message': str(e)
                })
                
                if not skip_errors:
                    break
        
        # Update import record
        import_record.success_count = success_count
        import_record.error_count = error_count
        import_record.warning_count = warning_count
        import_record.errors = errors
        import_record.warnings = warnings
        import_record.status = 'COMPLETED'
        import_record.save()
        
        return {
            'success': True,
            'success_count': success_count,
            'error_count': error_count,
            'warning_count': warning_count,
            'errors': errors,
            'warnings': warnings
        }
        
    except Exception as e:
        import_record.status = 'FAILED'
        import_record.errors = [{'row': 0, 'field': 'general', 'message': str(e)}]
        import_record.save()
        
        return {
            'success': False,
            'error': str(e)
        }

# API Testing Dashboard Views
@login_required
@user_passes_test(is_admin)
def api_testing_dashboard(request):
    """Main API testing dashboard"""
    context = {
        'total_collections': APICollection.objects.filter(created_by=request.user).count(),
        'total_requests': APIRequest.objects.filter(collection__created_by=request.user).count(),
        'total_tests': APITest.objects.filter(request__collection__created_by=request.user).count(),
        'total_environments': APIEnvironment.objects.filter(created_by=request.user).count(),
        'recent_results': APITestResult.objects.filter(
            test__request__collection__created_by=request.user
        ).order_by('-executed_at')[:10],
        'recent_suite_results': APITestSuiteResult.objects.filter(
            suite__collection__created_by=request.user
        ).order_by('-started_at')[:5],
        'active_automations': APIAutomation.objects.filter(
            test_suite__collection__created_by=request.user,
            is_active=True
        ).count(),
    }
    return render(request, 'dashboard/api_testing/dashboard.html', context)

@login_required
@user_passes_test(is_admin)
def api_collections_list(request):
    """API Collections management page"""
    if request.method == 'POST':
        # Handle collection creation
        try:
            collection = APICollection.objects.create(
                name=request.POST.get('name'),
                description=request.POST.get('description', ''),
                base_url=request.POST.get('base_url', ''),
                is_public=request.POST.get('is_public') == 'on',
                created_by=request.user
            )
            return JsonResponse({
                'success': True,
                'message': 'Collection created successfully',
                'collection_id': str(collection.id)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    collections = APICollection.objects.filter(created_by=request.user).order_by('-created_at')
    return render(request, 'dashboard/api_testing/collections.html', {'collections': collections})

@login_required
@user_passes_test(is_admin)
def api_collection_detail_view(request, collection_id):
    """Get collection details for AJAX requests"""
    try:
        collection = APICollection.objects.get(id=collection_id, created_by=request.user)
        return JsonResponse({
            'id': str(collection.id),
            'name': collection.name,
            'description': collection.description,
            'base_url': collection.base_url,
            'is_public': collection.is_public,
            'created_at': collection.created_at.isoformat(),
            'updated_at': collection.updated_at.isoformat()
        })
    except APICollection.DoesNotExist:
        return JsonResponse({'error': 'Collection not found'}, status=404)

@login_required
@user_passes_test(is_admin)
def api_collection_update_view(request, collection_id):
    """Update collection for AJAX requests"""
    try:
        collection = APICollection.objects.get(id=collection_id, created_by=request.user)
        if request.method == 'PUT':
            # Parse PUT data
            import json
            data = json.loads(request.body.decode('utf-8'))
            
            collection.name = data.get('name', collection.name)
            collection.description = data.get('description', collection.description)
            collection.base_url = data.get('base_url', collection.base_url)
            collection.is_public = data.get('is_public', collection.is_public)
            collection.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Collection updated successfully'
            })
    except APICollection.DoesNotExist:
        return JsonResponse({'error': 'Collection not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_collection_duplicate_view(request, collection_id):
    """Duplicate collection for AJAX requests"""
    try:
        collection = APICollection.objects.get(id=collection_id, created_by=request.user)
        
        # Create new collection
        new_collection = APICollection.objects.create(
            name=f"{collection.name} (Copy)",
            description=collection.description,
            base_url=collection.base_url,
            is_public=collection.is_public,
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Collection duplicated successfully',
            'new_collection_id': str(new_collection.id)
        })
    except APICollection.DoesNotExist:
        return JsonResponse({'error': 'Collection not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_collection_delete_view(request, collection_id):
    """Delete collection for AJAX requests"""
    try:
        collection = APICollection.objects.get(id=collection_id, created_by=request.user)
        collection.delete()
        return JsonResponse({
            'success': True,
            'message': 'Collection deleted successfully'
        })
    except APICollection.DoesNotExist:
        return JsonResponse({'error': 'Collection not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_collection_detail(request, collection_id):
    """API Collection detail page"""
    collection = get_object_or_404(APICollection, id=collection_id, created_by=request.user)
    return render(request, 'dashboard/api_testing/collection_detail.html', {'collection': collection})

@login_required
@user_passes_test(is_admin)
def api_environments_list(request):
    """API Environments management page"""
    if request.method == 'POST':
        # Handle environment creation
        try:
            import json
            variables = json.loads(request.POST.get('variables', '{}'))
            
            # If setting as default, unset other defaults
            if request.POST.get('is_default') == 'on':
                APIEnvironment.objects.filter(created_by=request.user, is_default=True).update(is_default=False)
            
            environment = APIEnvironment.objects.create(
                name=request.POST.get('name'),
                description=request.POST.get('description', ''),
                variables=variables,
                is_default=request.POST.get('is_default') == 'on',
                created_by=request.user
            )
            return JsonResponse({
                'success': True,
                'message': 'Environment created successfully',
                'environment_id': str(environment.id)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    environments = APIEnvironment.objects.filter(created_by=request.user).order_by('-created_at')
    return render(request, 'dashboard/api_testing/environments.html', {'environments': environments})

@login_required
@user_passes_test(is_admin)
def api_environment_detail_view(request, environment_id):
    """Get environment details for AJAX requests"""
    try:
        environment = APIEnvironment.objects.get(id=environment_id, created_by=request.user)
        return JsonResponse({
            'id': str(environment.id),
            'name': environment.name,
            'description': environment.description,
            'variables': environment.variables,
            'is_default': environment.is_default,
            'created_at': environment.created_at.isoformat(),
            'updated_at': environment.updated_at.isoformat()
        })
    except APIEnvironment.DoesNotExist:
        return JsonResponse({'error': 'Environment not found'}, status=404)

@login_required
@user_passes_test(is_admin)
def api_environment_update_view(request, environment_id):
    """Update environment for AJAX requests"""
    try:
        environment = APIEnvironment.objects.get(id=environment_id, created_by=request.user)
        if request.method == 'PUT':
            # Parse PUT data
            import json
            data = json.loads(request.body.decode('utf-8'))
            
            environment.name = data.get('name', environment.name)
            environment.description = data.get('description', environment.description)
            environment.variables = data.get('variables', environment.variables)
            environment.is_default = data.get('is_default', environment.is_default)
            
            # If setting as default, unset other defaults
            if environment.is_default:
                APIEnvironment.objects.filter(created_by=request.user, is_default=True).exclude(id=environment.id).update(is_default=False)
            
            environment.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Environment updated successfully'
            })
    except APIEnvironment.DoesNotExist:
        return JsonResponse({'error': 'Environment not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_environment_duplicate_view(request, environment_id):
    """Duplicate environment for AJAX requests"""
    try:
        environment = APIEnvironment.objects.get(id=environment_id, created_by=request.user)
        
        # Create new environment
        new_environment = APIEnvironment.objects.create(
            name=f"{environment.name} (Copy)",
            description=environment.description,
            variables=environment.variables,
            is_default=False,  # Don't duplicate default status
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Environment duplicated successfully',
            'new_environment_id': str(new_environment.id)
        })
    except APIEnvironment.DoesNotExist:
        return JsonResponse({'error': 'Environment not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_environment_set_default_view(request, environment_id):
    """Set environment as default for AJAX requests"""
    try:
        environment = APIEnvironment.objects.get(id=environment_id, created_by=request.user)
        
        # Unset other defaults
        APIEnvironment.objects.filter(created_by=request.user, is_default=True).update(is_default=False)
        
        # Set this environment as default
        environment.is_default = True
        environment.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Environment set as default successfully'
        })
    except APIEnvironment.DoesNotExist:
        return JsonResponse({'error': 'Environment not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_environment_delete_view(request, environment_id):
    """Delete environment for AJAX requests"""
    try:
        environment = APIEnvironment.objects.get(id=environment_id, created_by=request.user)
        environment.delete()
        return JsonResponse({
            'success': True,
            'message': 'Environment deleted successfully'
        })
    except APIEnvironment.DoesNotExist:
        return JsonResponse({'error': 'Environment not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@user_passes_test(is_admin)
def api_requests_list(request):
    """API Requests management page"""
    requests = APIRequest.objects.filter(collection__created_by=request.user).order_by('collection__name', 'order')
    return render(request, 'dashboard/api_testing/requests.html', {'requests': requests})

@login_required
@user_passes_test(is_admin)
def api_request_detail(request, request_id):
    """API Request detail page"""
    api_request = get_object_or_404(APIRequest, id=request_id, collection__created_by=request.user)
    return render(request, 'dashboard/api_testing/request_detail.html', {'api_request': api_request})

@login_required
@user_passes_test(is_admin)
def api_tests_list(request):
    """API Tests management page"""
    tests = APITest.objects.filter(request__collection__created_by=request.user).order_by('request__collection__name', 'request__order', 'name')
    return render(request, 'dashboard/api_testing/tests.html', {'tests': tests})

@login_required
@user_passes_test(is_admin)
def api_test_detail(request, test_id):
    """API Test detail page"""
    test = get_object_or_404(APITest, id=test_id, request__collection__created_by=request.user)
    return render(request, 'dashboard/api_testing/test_detail.html', {'test': test})

@login_required
@user_passes_test(is_admin)
def api_test_results_list(request):
    """API Test Results page"""
    results = APITestResult.objects.filter(
        test__request__collection__created_by=request.user
    ).order_by('-executed_at')
    return render(request, 'dashboard/api_testing/test_results.html', {'results': results})

@login_required
@user_passes_test(is_admin)
def api_test_suites_list(request):
    """API Test Suites management page"""
    suites = APITestSuite.objects.filter(collection__created_by=request.user).order_by('-created_at')
    return render(request, 'dashboard/api_testing/test_suites.html', {'suites': suites})

@login_required
@user_passes_test(is_admin)
def api_test_suite_detail(request, suite_id):
    """API Test Suite detail page"""
    suite = get_object_or_404(APITestSuite, id=suite_id, collection__created_by=request.user)
    return render(request, 'dashboard/api_testing/test_suite_detail.html', {'suite': suite})

@login_required
@user_passes_test(is_admin)
def api_automations_list(request):
    """API Automations management page"""
    automations = APIAutomation.objects.filter(test_suite__collection__created_by=request.user).order_by('-created_at')
    return render(request, 'dashboard/api_testing/automations.html', {'automations': automations})

@login_required
@user_passes_test(is_admin)
def api_automation_detail(request, automation_id):
    """API Automation detail page"""
    automation = get_object_or_404(APIAutomation, id=automation_id, test_suite__collection__created_by=request.user)
    return render(request, 'dashboard/api_testing/automation_detail.html', {'automation': automation})

@login_required
@user_passes_test(is_admin)
def api_testing_workspace(request):
    """Interactive API testing workspace"""
    collections = APICollection.objects.filter(created_by=request.user).order_by('name')
    environments = APIEnvironment.objects.filter(created_by=request.user).order_by('name')
    return render(request, 'dashboard/api_testing/workspace.html', {
        'collections': collections,
        'environments': environments
    })

@user_passes_test(is_admin)
def simple_api_workspace(request):
    """Simple API testing workspace - Postman-like interface"""
    collections = APICollection.objects.filter(created_by=request.user).order_by('name')
    return render(request, 'dashboard/api_testing/simple_workspace.html', {
        'collections': collections
    })

# ------------------------------
# Faculty Management Views
# ------------------------------
@login_required
@user_passes_test(is_admin)
def faculty_list(request):
    """Faculty management page"""
    # Get filter parameters
    search = request.GET.get('search', '')
    department = request.GET.get('department', '')
    status = request.GET.get('status', '')
    
    # Build queryset
    faculties = Faculty.objects.select_related('department_ref').all()
    
    if search:
        faculties = faculties.filter(
            Q(name__icontains=search) |
            Q(employee_id__icontains=search) |
            Q(apaar_faculty_id__icontains=search) |
            Q(email__icontains=search) |
            Q(phone_number__icontains=search)
        )
    
    if department:
        # Filter by department_ref if it's a UUID, otherwise by old department field
        try:
            from django.core.validators import UUIDValidator
            UUIDValidator()(department)
            faculties = faculties.filter(department_ref_id=department)
        except:
            faculties = faculties.filter(department=department)
    
    if status:
        faculties = faculties.filter(status=status)
    
    faculties = faculties.order_by('name')
    
    # Get statistics
    total_faculty = Faculty.objects.count()
    active_faculty = Faculty.objects.filter(status='ACTIVE', currently_associated=True).count()
    
    # Get departments for filter dropdown
    departments = Department.objects.filter(is_active=True, status='ACTIVE').order_by('name')
    
    context = {
        'faculties': faculties,
        'total_faculty': total_faculty,
        'active_faculty': active_faculty,
        'departments': departments,
        'department_choices': Faculty.DEPARTMENT_CHOICES,
        'status_choices': Faculty.STATUS_CHOICES,
    }
    return render(request, 'dashboard/faculty/list.html', context)

@login_required
@user_passes_test(is_admin)
def faculty_create(request):
    """Create new faculty member"""
    from faculty.forms import FacultyCreateForm
    
    if request.method == 'POST':
        form = FacultyCreateForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                faculty = form.save()
                messages.success(request, f'Faculty member {faculty.name} created successfully.')
                return redirect('dashboard:faculty_detail', faculty_id=faculty.id)
            except Exception as e:
                messages.error(request, f'Error creating faculty member: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = FacultyCreateForm()
    
    # Get departments for the form
    departments = Department.objects.filter(is_active=True, status='ACTIVE').order_by('name')
    
    context = {
        'form': form,
        'departments': departments,
        'title': 'Add New Faculty Member',
    }
    return render(request, 'dashboard/faculty/create.html', context)


@login_required
@user_passes_test(is_admin)
def faculty_update(request, faculty_id):
    """Update existing faculty member"""
    from faculty.forms import FacultyUpdateForm
    
    try:
        faculty = Faculty.objects.get(id=faculty_id)
    except Faculty.DoesNotExist:
        messages.error(request, 'Faculty member not found.')
        return redirect('dashboard:faculty_list')
    
    if request.method == 'POST':
        form = FacultyUpdateForm(request.POST, request.FILES, instance=faculty)
        if form.is_valid():
            try:
                faculty = form.save()
                messages.success(request, f'Faculty member {faculty.name} updated successfully.')
                return redirect('dashboard:faculty_detail', faculty_id=faculty.id)
            except Exception as e:
                messages.error(request, f'Error updating faculty member: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = FacultyUpdateForm(instance=faculty)
    
    # Get departments for the form
    departments = Department.objects.filter(is_active=True, status='ACTIVE').order_by('name')
    
    context = {
        'form': form,
        'faculty': faculty,
        'departments': departments,
        'title': f'Update {faculty.name}',
    }
    return render(request, 'dashboard/faculty/update.html', context)


@login_required
@user_passes_test(is_admin)
def faculty_detail(request, faculty_id):
    """Faculty detail page"""
    try:
        faculty = Faculty.objects.select_related('department_ref').get(id=faculty_id)
        context = {
            'faculty': faculty,
            'subjects': faculty.subjects.all(),
            'schedules': faculty.schedules.all(),
            'leaves': faculty.leaves.all(),
            'performance_records': faculty.performance_records.all(),
            'documents': faculty.documents.all(),
            'custom_fields': faculty.custom_field_values.all(),
        }
        return render(request, 'dashboard/faculty/detail.html', context)
    except Faculty.DoesNotExist:
        return render(request, 'dashboard/404.html', status=404)

@login_required
@user_passes_test(is_admin)
def custom_fields_list(request):
    """Custom fields management page"""
    from students.models import CustomField
    from django.utils import timezone
    
    custom_fields = CustomField.objects.all().order_by('order', 'name')
    
    # Calculate statistics
    active_fields_count = custom_fields.filter(is_active=True).count()
    required_fields_count = custom_fields.filter(required=True).count()
    field_types_count = custom_fields.values('field_type').distinct().count()
    
    context = {
        'custom_fields': custom_fields,
        'active_fields_count': active_fields_count,
        'required_fields_count': required_fields_count,
        'field_types_count': field_types_count,
        'now': timezone.now(),
    }
    return render(request, 'dashboard/custom_fields.html', context)


# ------------------------------
# Placements Dashboard Views
# ------------------------------
@login_required
@user_passes_test(is_admin)
def placements_dashboard(request):
    """Placements overview page."""
    from placements.models import Company, JobPosting, Application, PlacementDrive, InterviewRound, Offer
    context = {
        'companies_count': Company.objects.count(),
        'jobs_count': JobPosting.objects.count(),
        'applications_count': Application.objects.count(),
        'drives_count': PlacementDrive.objects.count(),
        'rounds_count': InterviewRound.objects.count(),
        'offers_count': Offer.objects.count(),
    }
    return render(request, 'dashboard/placements/dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def placements_companies(request):
    """List and create companies."""
    from placements.models import Company
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        website = request.POST.get('website', '').strip() or None
        industry = request.POST.get('industry', '').strip()
        description = request.POST.get('description', '').strip()
        headquarters = request.POST.get('headquarters', '').strip()
        contact_email = request.POST.get('contact_email', '').strip() or None
        contact_phone = request.POST.get('contact_phone', '').strip()
        if name:
            Company.objects.create(
                name=name,
                website=website,
                industry=industry,
                description=description,
                headquarters=headquarters,
                contact_email=contact_email,
                contact_phone=contact_phone,
            )
    search = request.GET.get('search', '').strip()
    companies = Company.objects.all()
    if search:
        companies = companies.filter(Q(name__icontains=search) | Q(industry__icontains=search))
    companies = companies.order_by('name')
    context = {'companies': companies}
    return render(request, 'dashboard/placements/companies.html', context)


@login_required
@user_passes_test(is_admin)
def placements_jobs(request):
    """List and create job postings."""
    from placements.models import Company, JobPosting, JobType, WorkMode
    if request.method == 'POST':
        company_id = request.POST.get('company_id')
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        location = request.POST.get('location', '').strip()
        work_mode = request.POST.get('work_mode') or WorkMode.ONSITE
        job_type = request.POST.get('job_type') or JobType.INTERNSHIP
        is_active = request.POST.get('is_active') == 'on'
        if title and company_id:
            try:
                company = Company.objects.get(id=company_id)
                JobPosting.objects.create(
                    company=company,
                    title=title,
                    description=description,
                    location=location,
                    work_mode=work_mode,
                    job_type=job_type,
                    is_active=is_active,
                    posted_by=request.user,
                )
            except Company.DoesNotExist:
                pass
    jobs = (
        JobPosting.objects.select_related('company')
        .all()
        .order_by('-created_at')
    )
    companies = Company.objects.order_by('name')
    context = {
        'jobs': jobs,
        'companies': companies,
        'job_types': [(c.value, c.label) for c in JobType],
        'work_modes': [(c.value, c.label) for c in WorkMode],
    }
    return render(request, 'dashboard/placements/jobs.html', context)


@login_required
@user_passes_test(is_admin)
def placements_applications(request):
    """List applications with simple filters."""
    from placements.models import Application, ApplicationStatus
    status = request.GET.get('status', '').strip()
    qs = Application.objects.select_related('student', 'job', 'job__company').all()
    if status:
        qs = qs.filter(status=status)
    applications = qs.order_by('-applied_at')
    context = {
        'applications': applications,
        'status_choices': [(c.value, c.label) for c in ApplicationStatus],
    }
    return render(request, 'dashboard/placements/applications.html', context)


@login_required
@user_passes_test(is_admin)
def placements_drives(request):
    """List and simple create for placement drives."""
    from placements.models import PlacementDrive, Company
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        company_name = request.POST.get('company', '').strip()
        drive_type = request.POST.get('drive_type') or 'CAMPUS'
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date') or None
        description = request.POST.get('description', '').strip()
        if title and company_name and start_date:
            company = Company.objects.filter(name__iexact=company_name).first()
            if company:
                PlacementDrive.objects.create(
                    company=company,
                    title=title,
                    description=description,
                    drive_type=drive_type,
                    start_date=start_date,
                    end_date=end_date or None,
                    created_by=request.user,
                )
    drives = PlacementDrive.objects.select_related('company').all().order_by('-start_date')
    companies = Company.objects.order_by('name')
    context = {
        'drives': drives,
        'companies': companies,
    }
    return render(request, 'dashboard/placements/drives.html', context)


@login_required
@user_passes_test(is_admin)
def placements_rounds(request):
    """List and create interview rounds linked to a drive."""
    from placements.models import InterviewRound, PlacementDrive, InterviewRoundType
    if request.method == 'POST':
        drive_id = request.POST.get('drive_id')
        name = request.POST.get('name', '').strip()
        round_type = request.POST.get('round_type') or InterviewRoundType.OTHER
        scheduled_at = request.POST.get('scheduled_at') or None
        location = request.POST.get('location', '').strip()
        instructions = request.POST.get('instructions', '').strip()
        if drive_id and name:
            try:
                drive = PlacementDrive.objects.get(id=drive_id)
                InterviewRound.objects.create(
                    drive=drive,
                    name=name,
                    round_type=round_type,
                    scheduled_at=scheduled_at or None,
                    location=location,
                    instructions=instructions,
                )
            except PlacementDrive.DoesNotExist:
                pass
    rounds = InterviewRound.objects.select_related('drive').all().order_by('-scheduled_at')
    drives = PlacementDrive.objects.order_by('-start_date')
    context = {
        'rounds': rounds,
        'drives': drives,
        'round_types': [(c.value, c.label) for c in InterviewRoundType],
    }
    return render(request, 'dashboard/placements/rounds.html', context)


@login_required
@user_passes_test(is_admin)
def placements_offers(request):
    """List and create offers for applications."""
    from placements.models import Offer, Application, OfferStatus
    if request.method == 'POST':
        application_id = request.POST.get('application_id')
        offered_role = request.POST.get('offered_role', '').strip()
        package_annual_ctc = request.POST.get('package_annual_ctc')
        joining_location = request.POST.get('joining_location', '').strip()
        joining_date = request.POST.get('joining_date') or None
        status = request.POST.get('status') or OfferStatus.PENDING
        if application_id and offered_role and package_annual_ctc:
            try:
                application = Application.objects.get(id=application_id)
                Offer.objects.update_or_create(
                    application=application,
                    defaults={
                        'offered_role': offered_role,
                        'package_annual_ctc': package_annual_ctc,
                        'joining_location': joining_location,
                        'joining_date': joining_date or None,
                        'status': status,
                    }
                )
            except Application.DoesNotExist:
                pass
    offers = Offer.objects.select_related('application', 'application__student', 'application__job', 'application__job__company').all().order_by('-offered_at')
    applications = Application.objects.select_related('student', 'job', 'job__company').all().order_by('-applied_at')
    context = {
        'offers': offers,
        'applications': applications,
        'status_choices': [(c.value, c.label) for c in OfferStatus],
    }
    return render(request, 'dashboard/placements/offers.html', context)

@login_required
@user_passes_test(is_admin)
def faculty_performance_stats(request):
    """Faculty performance statistics dashboard"""
    # Get total faculty count
    total_faculty = Faculty.objects.count()
    
    # Get performance statistics
    performance_stats = FacultyPerformance.objects.all().aggregate(
        total_performance_records=Count('id'),
        average_rating=Avg('overall_score')
    )
    
    # Get recent performance records
    recent_performance_records = FacultyPerformance.objects.order_by('-created_at')[:10]
    
    # Get department-wise statistics
    department_stats = Faculty.objects.values('department').annotate(
        count=Count('id')
    ).order_by('-count')
    
    context = {
        'total_faculty': total_faculty,
        'performance_stats': performance_stats,
        'recent_performance_records': recent_performance_records,
        'department_stats': department_stats,
    }
    return render(request, 'dashboard/faculty/performance_stats.html', context)

@login_required
@user_passes_test(is_admin)
def faculty_leave_stats(request):
    """Faculty leave statistics dashboard"""
    # Get total faculty count
    total_faculty = Faculty.objects.count()
    
    # Get leave statistics
    leave_stats = FacultyLeave.objects.all().aggregate(
        total_leaves=Count('id'),
        approved_leaves=Count('id', filter=Q(status='APPROVED')),
        pending_leaves=Count('id', filter=Q(status='PENDING')),
        rejected_leaves=Count('id', filter=Q(status='REJECTED')),
    )
    
    # Get recent leaves
    recent_leaves = FacultyLeave.objects.order_by('-created_at')[:10]
    
    # Get leave type statistics
    leave_type_stats = FacultyLeave.objects.values('leave_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    context = {
        'total_faculty': total_faculty,
        'leave_stats': leave_stats,
        'recent_leaves': recent_leaves,
        'leave_type_stats': leave_type_stats,
    }
    return render(request, 'dashboard/faculty/leave_stats.html', context)

@login_required
@user_passes_test(is_admin)
def faculty_document_list(request):
    """Faculty document management page"""
    documents = FacultyDocument.objects.all().order_by('-created_at')
    
    context = {
        'documents': documents,
    }
    return render(request, 'dashboard/faculty/documents.html', context)

@login_required
@user_passes_test(is_admin)
def faculty_custom_field_create(request):
    """Create a new faculty custom field"""
    if request.method == 'POST':
        try:
            # Get form data
            name = request.POST.get('name')
            label = request.POST.get('label')
            field_type = request.POST.get('field_type')
            required = request.POST.get('required') == 'on'
            is_active = request.POST.get('is_active') == 'on'
            default_value = request.POST.get('default_value', '')
            choices = request.POST.get('choices', '')
            help_text = request.POST.get('help_text', '')
            order = request.POST.get('order', 0)
            
            # Validate required fields
            if not all([name, label, field_type]):
                return JsonResponse({'success': False, 'error': 'Name, label, and field type are required'})
            
            # Check if name already exists
            if FacultyCustomField.objects.filter(name=name).exists():
                return JsonResponse({'success': False, 'error': 'A field with this name already exists'})
            
            # Create the custom field
            custom_field = FacultyCustomField.objects.create(
                name=name,
                label=label,
                field_type=field_type,
                required=required,
                is_active=is_active,
                default_value=default_value,
                choices=choices,
                help_text=help_text,
                order=order
            )
            
            return JsonResponse({'success': True, 'message': 'Custom field created successfully'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
@user_passes_test(is_admin)
def faculty_custom_field_update(request, field_id):
    """Update an existing faculty custom field"""
    if request.method == 'POST':
        try:
            # Get the custom field
            try:
                custom_field = FacultyCustomField.objects.get(id=field_id)
            except FacultyCustomField.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Custom field not found'})
            
            # Get form data
            name = request.POST.get('name')
            label = request.POST.get('label')
            field_type = request.POST.get('field_type')
            required = request.POST.get('required') == 'on'
            is_active = request.POST.get('is_active') == 'on'
            default_value = request.POST.get('default_value', '')
            choices = request.POST.get('choices', '')
            help_text = request.POST.get('help_text', '')
            order = request.POST.get('order', 0)
            
            # Validate required fields
            if not all([name, label, field_type]):
                return JsonResponse({'success': False, 'error': 'Name, label, and field type are required'})
            
            # Check if name already exists (excluding current field)
            if FacultyCustomField.objects.filter(name=name).exclude(id=field_id).exists():
                return JsonResponse({'success': False, 'error': 'A field with this name already exists'})
            
            # Update the custom field
            custom_field.name = name
            custom_field.label = label
            custom_field.field_type = field_type
            custom_field.required = required
            custom_field.is_active = is_active
            custom_field.default_value = default_value
            custom_field.choices = choices
            custom_field.help_text = help_text
            custom_field.order = order
            custom_field.save()
            
            return JsonResponse({'success': True, 'message': 'Custom field updated successfully'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
@user_passes_test(is_admin)
def faculty_custom_field_delete(request, field_id):
    """Delete a faculty custom field"""
    if request.method == 'POST':
        try:
            # Get the custom field
            try:
                custom_field = FacultyCustomField.objects.get(id=field_id)
            except FacultyCustomField.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Custom field not found'})
            
            # Delete the custom field
            custom_field.delete()
            
            return JsonResponse({'success': True, 'message': 'Custom field deleted successfully'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
@user_passes_test(is_admin)
def faculty_custom_fields_list(request):
    """Faculty custom fields management page"""
    custom_fields = FacultyCustomField.objects.all().order_by('order', 'name')
    
    # Calculate statistics
    active_fields_count = custom_fields.filter(is_active=True).count()
    required_fields_count = custom_fields.filter(required=True).count()
    field_types_count = custom_fields.values('field_type').distinct().count()
    
    context = {
        'custom_fields': custom_fields,
        'active_fields_count': active_fields_count,
        'required_fields_count': required_fields_count,
        'field_types_count': field_types_count,
        'now': timezone.now(),
    }
    return render(request, 'dashboard/faculty/custom_fields.html', context)


# ============================================================================
# ACADEMICS DASHBOARD VIEWS
# ============================================================================

@login_required
@user_passes_test(is_admin)
def academics_dashboard(request):
    """Main academics dashboard with statistics"""
    context = {
        'total_courses': Course.objects.count(),
        'active_courses': Course.objects.filter(status='ACTIVE').count(),
        'total_syllabi': Syllabus.objects.count(),
        'approved_syllabi': Syllabus.objects.filter(status='APPROVED').count(),
        'total_timetables': Timetable.objects.filter(is_active=True).count(),
        'total_enrollments': CourseEnrollment.objects.count(),
        'active_enrollments': CourseEnrollment.objects.filter(status='ENROLLED').count(),
        'total_calendar_events': AcademicCalendar.objects.count(),
        'upcoming_events': AcademicCalendar.objects.filter(start_date__gte=timezone.now().date()).count(),
        'recent_courses': Course.objects.order_by('-created_at')[:5],
        'recent_syllabi': Syllabus.objects.order_by('-created_at')[:5],
        'recent_enrollments': CourseEnrollment.objects.order_by('-created_at')[:5],
        'recent_events': AcademicCalendar.objects.order_by('-created_at')[:5],
    }
    return render(request, 'dashboard/academics/dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def academics_courses_list(request):
    """Courses management page"""
    courses = Course.objects.all().order_by('code')
    
    # Filtering
    level_filter = request.GET.get('level')
    status_filter = request.GET.get('status')
    
    if level_filter:
        courses = courses.filter(level=level_filter)
    if status_filter:
        courses = courses.filter(status=status_filter)
    
    # Statistics
    context = {
        'courses': courses,
        'total_courses': Course.objects.count(),
        'active_courses': Course.objects.filter(status='ACTIVE').count(),
        'level_choices': Course.COURSE_LEVELS,
        'status_choices': Course.COURSE_STATUS,
        'current_filters': {
            'level': level_filter,
            'status': status_filter,
        }
    }
    return render(request, 'dashboard/academics/courses.html', context)


@login_required
@user_passes_test(is_admin)
def academics_course_detail(request, course_id):
    """Course detail page"""
    course = get_object_or_404(Course, id=course_id)
    
    # Get related data
    syllabus = getattr(course, 'syllabus', None)
    timetables = course.timetables.filter(is_active=True).order_by('day_of_week', 'start_time')
    enrollments = course.enrollments.all().order_by('-academic_year', '-semester')
    
    context = {
        'course': course,
        'syllabus': syllabus,
        'timetables': timetables,
        'enrollments': enrollments,
        'enrollment_count': enrollments.count(),
        'active_enrollments': enrollments.filter(status='ENROLLED').count(),
    }
    return render(request, 'dashboard/academics/course_detail.html', context)


@login_required
@user_passes_test(is_admin)
def academics_syllabi_list(request):
    """Syllabi management page"""
    syllabi = Syllabus.objects.all().order_by('-academic_year', '-semester', 'course__code')
    
    # Filtering
    status_filter = request.GET.get('status')
    academic_year_filter = request.GET.get('academic_year')
    
    if status_filter:
        syllabi = syllabi.filter(status=status_filter)
    if academic_year_filter:
        syllabi = syllabi.filter(academic_year=academic_year_filter)
    
    context = {
        'syllabi': syllabi,
        'total_syllabi': Syllabus.objects.count(),
        'approved_syllabi': Syllabus.objects.filter(status='APPROVED').count(),
        'status_choices': Syllabus.SYLLABUS_STATUS,
        'academic_years': Syllabus.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year'),
        'current_filters': {
            'status': status_filter,
            'academic_year': academic_year_filter,
        }
    }
    return render(request, 'dashboard/academics/syllabi.html', context)


@login_required
@user_passes_test(is_admin)
def academics_syllabus_detail(request, syllabus_id):
    """Syllabus detail page"""
    syllabus = get_object_or_404(Syllabus, id=syllabus_id)
    topics = syllabus.topics.all().order_by('week_number', 'order')
    
    context = {
        'syllabus': syllabus,
        'topics': topics,
        'topics_count': topics.count(),
        'total_hours': sum(topic.duration_hours for topic in topics),
    }
    return render(request, 'dashboard/academics/syllabus_detail.html', context)


@login_required
@user_passes_test(is_admin)
def academics_timetables_list(request):
    """Timetables management page"""
    timetables = Timetable.objects.all().order_by('academic_year', 'semester', 'day_of_week', 'start_time')
    
    # Filtering
    academic_year_filter = request.GET.get('academic_year')
    semester_filter = request.GET.get('semester')
    day_filter = request.GET.get('day')
    
    if academic_year_filter:
        timetables = timetables.filter(academic_year=academic_year_filter)
    if semester_filter:
        timetables = timetables.filter(semester=semester_filter)
    if day_filter:
        timetables = timetables.filter(day_of_week=day_filter)
    
    context = {
        'timetables': timetables,
        'total_timetables': Timetable.objects.count(),
        'active_timetables': Timetable.objects.filter(is_active=True).count(),
        'academic_years': Timetable.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year'),
        'semesters': Timetable.objects.values_list('semester', flat=True).distinct().order_by('semester'),
        'days_of_week': Timetable.DAYS_OF_WEEK,
        'current_filters': {
            'academic_year': academic_year_filter,
            'semester': semester_filter,
            'day': day_filter,
        }
    }
    return render(request, 'dashboard/academics/timetables.html', context)


@login_required
@user_passes_test(is_admin)
def academics_timetable_detail(request, timetable_id):
    """Timetable detail page"""
    timetable = get_object_or_404(Timetable, id=timetable_id)
    
    # Check for conflicts
    conflicts = Timetable.objects.filter(
        academic_year=timetable.academic_year,
        semester=timetable.semester,
        day_of_week=timetable.day_of_week,
        room=timetable.room,
        is_active=True
    ).exclude(id=timetable.id)
    
    context = {
        'timetable': timetable,
        'conflicts': conflicts,
        'has_conflicts': conflicts.exists(),
        'duration_minutes': timetable.get_duration_minutes(),
    }
    return render(request, 'dashboard/academics/timetable_detail.html', context)


@login_required
@user_passes_test(is_admin)
def academics_enrollments_list(request):
    """Enrollments management page"""
    enrollments = CourseEnrollment.objects.all().order_by('-academic_year', '-semester', 'course__code')
    
    # Filtering
    status_filter = request.GET.get('status')
    academic_year_filter = request.GET.get('academic_year')
    semester_filter = request.GET.get('semester')
    
    if status_filter:
        enrollments = enrollments.filter(status=status_filter)
    if academic_year_filter:
        enrollments = enrollments.filter(academic_year=academic_year_filter)
    if semester_filter:
        enrollments = enrollments.filter(semester=semester_filter)
    
    context = {
        'enrollments': enrollments,
        'total_enrollments': CourseEnrollment.objects.count(),
        'active_enrollments': CourseEnrollment.objects.filter(status='ENROLLED').count(),
        'status_choices': CourseEnrollment.ENROLLMENT_STATUS,
        'academic_years': CourseEnrollment.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year'),
        'semesters': CourseEnrollment.objects.values_list('semester', flat=True).distinct().order_by('semester'),
        'current_filters': {
            'status': status_filter,
            'academic_year': academic_year_filter,
            'semester': semester_filter,
        }
    }
    return render(request, 'dashboard/academics/enrollments.html', context)


@login_required
@user_passes_test(is_admin)
def academics_enrollment_detail(request, enrollment_id):
    """Enrollment detail page"""
    enrollment = get_object_or_404(CourseEnrollment, id=enrollment_id)
    
    context = {
        'enrollment': enrollment,
        'student': enrollment.student,
        'course': enrollment.course,
    }
    return render(request, 'dashboard/academics/enrollment_detail.html', context)


@login_required
@user_passes_test(is_admin)
def academics_calendar_list(request):
    """Academic calendar management page"""
    events = AcademicCalendar.objects.all().order_by('start_date')
    
    # Filtering
    event_type_filter = request.GET.get('event_type')
    academic_year_filter = request.GET.get('academic_year')
    month_filter = request.GET.get('month')
    
    if event_type_filter:
        events = events.filter(event_type=event_type_filter)
    if academic_year_filter:
        events = events.filter(academic_year=academic_year_filter)
    if month_filter:
        events = events.filter(start_date__month=month_filter)
    
    context = {
        'events': events,
        'total_events': AcademicCalendar.objects.count(),
        'upcoming_events': AcademicCalendar.objects.filter(start_date__gte=timezone.now().date()).count(),
        'event_type_choices': AcademicCalendar.EVENT_TYPE,
        'academic_years': AcademicCalendar.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year'),
        'months': [
            (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
            (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
            (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
        ],
        'current_filters': {
            'event_type': event_type_filter,
            'academic_year': academic_year_filter,
            'month': month_filter,
        }
    }
    return render(request, 'dashboard/academics/calendar.html', context)


@login_required
@user_passes_test(is_admin)
def academics_calendar_detail(request, event_id):
    """Academic calendar event detail page"""
    event = get_object_or_404(AcademicCalendar, id=event_id)
    
    context = {
        'event': event,
        'duration_days': (event.end_date - event.start_date).days + 1,
    }
    return render(request, 'dashboard/academics/calendar_detail.html', context)

# ============================================================================
# ACADEMICS API ENDPOINTS DASHBOARD VIEWS
# ============================================================================

@login_required
@user_passes_test(is_admin)
def academics_api_endpoints(request):
    """Academics API endpoints overview and testing page"""
    context = {
        'api_base_url': '/api/v1/academics/',
        'endpoints': [
            {
                'name': 'Courses API',
                'url': '/api/v1/academics/courses/',
                'methods': ['GET', 'POST', 'PUT', 'PATCH', 'DELETE'],
                'description': 'Manage academic courses',
                'example_request': {
                    'method': 'POST',
                    'url': '/api/v1/academics/courses/',
                    'data': {
                        'code': 'CS101',
                        'title': 'Introduction to Computer Science',
                        'description': 'Basic concepts of computer science',
                        'credits': 3,
                        'level': 'UNDERGRADUATE',
                        'status': 'ACTIVE'
                    }
                }
            },
            {
                'name': 'Syllabi API',
                'url': '/api/v1/academics/syllabi/',
                'methods': ['GET', 'POST', 'PUT', 'PATCH', 'DELETE'],
                'description': 'Manage course syllabi and topics',
                'example_request': {
                    'method': 'POST',
                    'url': '/api/v1/academics/syllabi/',
                    'data': {
                        'course': 1,
                        'academic_year': '2024-2025',
                        'semester': 'FALL',
                        'learning_objectives': 'Understand basic programming concepts',
                        'status': 'DRAFT'
                    }
                }
            },
            {
                'name': 'Syllabus Topics API',
                'url': '/api/v1/academics/syllabus-topics/',
                'methods': ['GET', 'POST', 'PUT', 'PATCH', 'DELETE'],
                'description': 'Manage individual syllabus topics',
                'example_request': {
                    'method': 'POST',
                    'url': '/api/v1/academics/syllabus-topics/',
                    'data': {
                        'syllabus': 1,
                        'title': 'Variables and Data Types',
                        'description': 'Introduction to variables and basic data types',
                        'week_number': 1,
                        'hours_allocated': 2
                    }
                }
            },
            {
                'name': 'Timetables API',
                'url': '/api/v1/academics/timetables/',
                'methods': ['GET', 'POST', 'PUT', 'PATCH', 'DELETE'],
                'description': 'Manage class timetables and schedules',
                'example_request': {
                    'method': 'POST',
                    'url': '/api/v1/academics/timetables/',
                    'data': {
                        'course': 1,
                        'academic_year': '2024-2025',
                        'semester': 'FALL',
                        'day_of_week': 'MONDAY',
                        'start_time': '09:00:00',
                        'end_time': '10:30:00',
                        'room': 'Room 101',
                        'faculty': 1
                    }
                }
            },
            {
                'name': 'Enrollments API',
                'url': '/api/v1/academics/enrollments/',
                'methods': ['GET', 'POST', 'PUT', 'PATCH', 'DELETE'],
                'description': 'Manage student course enrollments',
                'example_request': {
                    'method': 'POST',
                    'url': '/api/v1/academics/enrollments/',
                    'data': {
                        'student': 1,
                        'course': 1,
                        'academic_year': '2024-2025',
                        'semester': 'FALL',
                        'enrollment_date': '2024-08-15',
                        'status': 'ENROLLED'
                    }
                }
            },
            {
                'name': 'Academic Calendar API',
                'url': '/api/v1/academics/academic-calendar/',
                'methods': ['GET', 'POST', 'PUT', 'PATCH', 'DELETE'],
                'description': 'Manage academic calendar events',
                'example_request': {
                    'method': 'POST',
                    'url': '/api/v1/academics/academic-calendar/',
                    'data': {
                        'title': 'Fall Semester Begins',
                        'event_type': 'ACADEMIC',
                        'start_date': '2024-08-26',
                        'end_date': '2024-08-26',
                        'academic_year': '2024-2025',
                        'description': 'First day of Fall semester classes'
                    }
                }
            }
        ],
        'authentication': {
            'type': 'JWT Token',
            'header': 'Authorization: Bearer <your_token>',
            'token_endpoint': '/api/auth/token/',
            'refresh_endpoint': '/api/auth/token/refresh/'
        },
        'pagination': {
            'default_page_size': 10,
            'max_page_size': 100,
            'example': '?page=1&page_size=20'
        },
        'filtering': {
            'search': '?search=computer',
            'ordering': '?ordering=title',
            'filtering': '?status=ACTIVE&level=UNDERGRADUATE'
        }
    }
    return render(request, 'dashboard/academics/api_endpoints.html', context)


@login_required
@user_passes_test(is_admin)
def academics_api_test(request):
    """API testing interface for academics endpoints"""
    context = {
        'api_base_url': '/api/v1/academics/',
        'test_endpoints': [
            {
                'name': 'Test Courses API',
                'url': '/api/v1/academics/courses/',
                'method': 'GET',
                'description': 'Test retrieving all courses'
            },
            {
                'name': 'Test Single Course',
                'url': '/api/v1/academics/courses/1/',
                'method': 'GET',
                'description': 'Test retrieving a specific course'
            },
            {
                'name': 'Test Syllabi API',
                'url': '/api/v1/academics/syllabi/',
                'method': 'GET',
                'description': 'Test retrieving all syllabi'
            },
            {
                'name': 'Test Timetables API',
                'url': '/api/v1/academics/timetables/',
                'method': 'GET',
                'description': 'Test retrieving all timetables'
            },
            {
                'name': 'Test Enrollments API',
                'url': '/api/v1/academics/enrollments/',
                'method': 'GET',
                'description': 'Test retrieving all enrollments'
            },
            {
                'name': 'Test Academic Calendar API',
                'url': '/api/v1/academics/academic-calendar/',
                'method': 'GET',
                'description': 'Test retrieving all calendar events'
            }
        ]
    }
    return render(request, 'dashboard/academics/api_test.html', context)


@login_required
@user_passes_test(is_admin)
def enrollment_dashboard(request):
    """Main enrollment management dashboard"""
    context = {
        'total_enrollment_rules': EnrollmentRule.objects.count(),
        'total_course_assignments': CourseAssignment.objects.count(),
        'total_faculty_assignments': FacultyAssignment.objects.count(),
        'total_enrollment_plans': StudentEnrollmentPlan.objects.count(),
        'total_enrollment_requests': EnrollmentRequest.objects.count(),
        'total_waitlist_entries': WaitlistEntry.objects.filter(is_active=True).count(),
        'pending_requests': EnrollmentRequest.objects.filter(status='PENDING').count(),
        'active_enrollments': CourseEnrollment.objects.filter(status='ENROLLED').count(),
        'recent_requests': EnrollmentRequest.objects.order_by('-request_date')[:5],
        'recent_plans': StudentEnrollmentPlan.objects.order_by('-created_at')[:5],
    }
    return render(request, 'dashboard/enrollment/dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def enrollment_rules_list(request):
    """List all enrollment rules"""
    rules = EnrollmentRule.objects.all().order_by('-academic_year', '-semester', 'rule_type')
    return render(request, 'dashboard/enrollment/rules.html', {'rules': rules})


@login_required
@user_passes_test(is_admin)
def enrollment_rule_detail(request, rule_id):
    """Detail view for enrollment rule"""
    rule = get_object_or_404(EnrollmentRule, pk=rule_id)
    return render(request, 'dashboard/enrollment/rule_detail.html', {'rule': rule})


@login_required
@user_passes_test(is_admin)
def course_assignments_list(request):
    """List all course assignments"""
    assignments = CourseAssignment.objects.all().order_by('-academic_year', '-semester', 'course__code')
    return render(request, 'dashboard/enrollment/course_assignments.html', {'assignments': assignments})


@login_required
@user_passes_test(is_admin)
def course_assignment_detail(request, assignment_id):
    """Detail view for course assignment"""
    assignment = get_object_or_404(CourseAssignment, pk=assignment_id)
    return render(request, 'dashboard/enrollment/course_assignment_detail.html', {'assignment': assignment})


@login_required
@user_passes_test(is_admin)
def faculty_assignments_list(request):
    """List all faculty assignments"""
    assignments = FacultyAssignment.objects.all().order_by('-assignment_date', 'faculty__first_name')
    return render(request, 'dashboard/enrollment/faculty_assignments.html', {'assignments': assignments})


@login_required
@user_passes_test(is_admin)
def faculty_assignment_detail(request, assignment_id):
    """Detail view for faculty assignment"""
    assignment = get_object_or_404(FacultyAssignment, pk=assignment_id)
    return render(request, 'dashboard/enrollment/faculty_assignment_detail.html', {'assignment': assignment})


@login_required
@user_passes_test(is_admin)
def enrollment_plans_list(request):
    """List all student enrollment plans"""
    plans = StudentEnrollmentPlan.objects.all().order_by('-academic_year', '-semester', 'student__roll_number')
    return render(request, 'dashboard/enrollment/enrollment_plans.html', {'plans': plans})


@login_required
@user_passes_test(is_admin)
def enrollment_plan_detail(request, plan_id):
    """Detail view for enrollment plan"""
    plan = get_object_or_404(StudentEnrollmentPlan, pk=plan_id)
    return render(request, 'dashboard/enrollment/enrollment_plan_detail.html', {'plan': plan})


@login_required
@user_passes_test(is_admin)
def enrollment_requests_list(request):
    """List all enrollment requests"""
    requests = EnrollmentRequest.objects.all().order_by('-request_date')
    return render(request, 'dashboard/enrollment/enrollment_requests.html', {'requests': requests})


@login_required
@user_passes_test(is_admin)
def enrollment_request_detail(request, request_id):
    """Detail view for enrollment request"""
    enrollment_request = get_object_or_404(EnrollmentRequest, pk=request_id)
    return render(request, 'dashboard/enrollment/enrollment_request_detail.html', {'enrollment_request': enrollment_request})


@login_required
@user_passes_test(is_admin)
def waitlist_entries_list(request):
    """List all waitlist entries"""
    entries = WaitlistEntry.objects.all().order_by('position', 'added_date')
    return render(request, 'dashboard/enrollment/waitlist_entries.html', {'entries': entries})


@login_required
@user_passes_test(is_admin)
def waitlist_entry_detail(request, entry_id):
    """Detail view for waitlist entry"""
    entry = get_object_or_404(WaitlistEntry, pk=entry_id)
    return render(request, 'dashboard/enrollment/waitlist_entry_detail.html', {'entry': entry})


@login_required
@user_passes_test(is_admin)
def departments_list(request):
    """List all departments"""
    departments = Department.objects.all().order_by('code', 'name')
    return render(request, 'dashboard/enrollment/departments.html', {'departments': departments})


@login_required
@user_passes_test(is_admin)
def department_detail(request, department_id):
    """Detail view for department"""
    department = get_object_or_404(Department, pk=department_id)
    return render(request, 'dashboard/enrollment/department_detail.html', {'department': department})


@login_required
@user_passes_test(is_admin)
def academic_programs_list(request):
    """List all academic programs"""
    programs = AcademicProgram.objects.all().order_by('level', 'code', 'name')
    return render(request, 'dashboard/enrollment/academic_programs.html', {'programs': programs})


@login_required
@user_passes_test(is_admin)
def academic_program_detail(request, program_id):
    """Detail view for academic program"""
    program = get_object_or_404(AcademicProgram, pk=program_id)
    return render(request, 'dashboard/enrollment/academic_program_detail.html', {'program': program})


@login_required
@user_passes_test(is_admin)
def course_sections_list(request):
    """List all course sections"""
    sections = CourseSection.objects.all().order_by('course__code', 'section_number', 'academic_year', 'semester')
    return render(request, 'dashboard/enrollment/course_sections.html', {'sections': sections})


@login_required
@user_passes_test(is_admin)
def course_section_detail(request, section_id):
    """Detail view for course section"""
    section = get_object_or_404(CourseSection, pk=section_id)
    return render(request, 'dashboard/enrollment/course_section_detail.html', {'section': section})


# ============================================================================
# ATTENDANCE DASHBOARD VIEWS
# ============================================================================

@login_required
@user_passes_test(is_admin)
def attendance_sessions(request):
    """List attendance sessions with filters and quick actions."""
    sessions = AttendanceSession.objects.select_related('course_section').order_by('-date', 'start_time')

    # Filters
    section_id = request.GET.get('section_id')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    is_cancelled = request.GET.get('is_cancelled')

    if section_id:
        sessions = sessions.filter(course_section_id=section_id)
    if date_from:
        sessions = sessions.filter(date__gte=date_from)
    if date_to:
        sessions = sessions.filter(date__lte=date_to)
    if is_cancelled in ['true', 'false']:
        sessions = sessions.filter(is_cancelled=(is_cancelled == 'true'))

    context = {
        'sessions': sessions[:200],
        'sections': CourseSection.objects.all().order_by('course__code', 'section_number'),
    }
    return render(request, 'dashboard/attendance/sessions.html', context)


@login_required
@user_passes_test(is_admin)
def attendance_session_detail(request, session_id):
    """Show a single session and all attendance records."""
    session = get_object_or_404(AttendanceSession.objects.select_related('course_section'), pk=session_id)
    records = AttendanceRecord.objects.filter(session=session).select_related('student').order_by('student__roll_number')
    context = {
        'session': session,
        'records': records,
        'status_choices': AttendanceRecord.STATUS_CHOICES,
    }
    return render(request, 'dashboard/attendance/session_detail.html', context)


@login_required
@user_passes_test(is_admin)
def attendance_mark(request, session_id):
    """Handle marking attendance for a session (AJAX or form POST)."""
    session = get_object_or_404(AttendanceSession, pk=session_id)
    if request.method == 'POST':
        try:
            student_id = request.POST.get('student_id')
            status_value = request.POST.get('status')
            remarks = request.POST.get('remarks', '')

            if status_value not in dict(AttendanceRecord.STATUS_CHOICES):
                return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)

            record, _ = AttendanceRecord.objects.get_or_create(session=session, student_id=student_id)
            record.status = status_value
            record.remarks = remarks
            record.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    # GET fallthrough
    return attendance_session_detail(request, session_id)


@login_required
@user_passes_test(is_admin)
def attendance_generate_sessions(request):
    """Form to generate sessions for a given date range from active timetables."""
    message = None
    error = None
    if request.method == 'POST':
        try:
            start = request.POST.get('start')
            end = request.POST.get('end')
            section_id = request.POST.get('section_id') or None

            # Use the same logic as the management command
            from datetime import date, timedelta
            start_date = date.fromisoformat(start)
            end_date = date.fromisoformat(end)
            if end_date < start_date:
                raise ValueError('End date must be after start date')

            days_map = {'MON': 0, 'TUE': 1, 'WED': 2, 'THU': 3, 'FRI': 4, 'SAT': 5, 'SUN': 6}
            timetables = Timetable.objects.filter(is_active=True)
            if section_id:
                timetables = timetables.filter(course_section_id=section_id)

            by_weekday = {}
            for t in timetables.select_related('course_section'):
                by_weekday.setdefault(days_map[t.day_of_week], []).append(t)

            created_count = 0
            current = start_date
            while current <= end_date:
                weekday = current.weekday()
                for t in by_weekday.get(weekday, []):
                    _, created = AttendanceSession.objects.get_or_create(
                        course_section=t.course_section,
                        date=current,
                        start_time=t.start_time,
                        defaults={'end_time': t.end_time, 'room': t.room, 'timetable': t}
                    )
                    if created:
                        created_count += 1
                current += timedelta(days=1)

            message = f'Created {created_count} attendance sessions.'
        except Exception as exc:
            error = str(exc)

    context = {
        'sections': CourseSection.objects.all().order_by('course__code', 'section_number'),
        'message': message,
        'error': error,
    }
    return render(request, 'dashboard/attendance/generate.html', context)


# ============================================================================
# EXAM MANAGEMENT DASHBOARD VIEWS
# ============================================================================

@login_required
@user_passes_test(is_admin)
def exams_dashboard(request):
    """Main exam management dashboard."""
    from exams.models import (
        ExamSession, ExamSchedule, ExamRegistration, HallTicket,
        ExamAttendance, ExamResult, StudentDue, ExamViolation,
        ExamRoom, ExamStaffAssignment, ExamRoomAllocation
    )
    from students.models import Student
    from academics.models import Course
    from faculty.models import Faculty
    from django.utils import timezone
    from datetime import timedelta
    
    # Get current date and time
    now = timezone.now()
    today = now.date()
    
    # Calculate statistics
    active_sessions_count = ExamSession.objects.filter(
        status__in=['PUBLISHED', 'ONGOING'],
        start_date__lte=today,
        end_date__gte=today
    ).count()
    
    upcoming_exams_count = ExamSchedule.objects.filter(
        exam_date__gte=today,
        status='SCHEDULED'
    ).count()
    
    total_registrations_count = ExamRegistration.objects.count()
    
    pending_approvals_count = ExamRegistration.objects.filter(
        status='PENDING'
    ).count()
    
    # Get upcoming exams for display
    upcoming_exams = ExamSchedule.objects.filter(
        exam_date__gte=today,
        status='SCHEDULED'
    ).select_related('course', 'exam_session').order_by('exam_date', 'start_time')[:5]
    
    # Get recent activities (simplified for now)
    recent_activities = []
    
    # Quick statistics
    total_students_count = Student.objects.count()
    total_courses_count = Course.objects.count()
    total_rooms_count = ExamRoom.objects.count()
    total_faculty_count = Faculty.objects.count()
    total_violations_count = ExamViolation.objects.count()
    total_results_count = ExamResult.objects.count()
    
    context = {
        'active_sessions_count': active_sessions_count,
        'upcoming_exams_count': upcoming_exams_count,
        'total_registrations_count': total_registrations_count,
        'pending_approvals_count': pending_approvals_count,
        'upcoming_exams': upcoming_exams,
        'recent_activities': recent_activities,
        'total_students_count': total_students_count,
        'total_courses_count': total_courses_count,
        'total_rooms_count': total_rooms_count,
        'total_faculty_count': total_faculty_count,
        'total_violations_count': total_violations_count,
        'total_results_count': total_results_count,
    }
    
    return render(request, 'dashboard/exams/dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def exams_sessions_list(request):
    """List all exam sessions."""
    from exams.models import ExamSession
    
    if request.method == 'POST':
        # Handle session creation
        try:
            from datetime import datetime
            from django.utils import timezone
            
            # Parse datetime strings
            registration_start = datetime.fromisoformat(request.POST['registration_start'].replace('Z', '+00:00'))
            registration_end = datetime.fromisoformat(request.POST['registration_end'].replace('Z', '+00:00'))
            
            # Make timezone-aware
            registration_start = timezone.make_aware(registration_start)
            registration_end = timezone.make_aware(registration_end)
            
            session = ExamSession.objects.create(
                name=request.POST['name'],
                session_type=request.POST['session_type'],
                academic_year=request.POST['academic_year'],
                semester=int(request.POST['semester']),
                start_date=request.POST['start_date'],
                end_date=request.POST['end_date'],
                registration_start=registration_start,
                registration_end=registration_end,
                description=request.POST.get('description', ''),
                status='DRAFT'
            )
            
            messages.success(request, f'Exam session "{session.name}" created successfully!')
            return redirect('dashboard:exams_sessions')
            
        except Exception as e:
            messages.error(request, f'Error creating exam session: {str(e)}')
    
    exam_sessions = ExamSession.objects.all().order_by('-academic_year', '-semester', '-start_date')
    
    context = {
        'exam_sessions': exam_sessions,
    }
    
    return render(request, 'dashboard/exams/sessions_list.html', context)


@login_required
@user_passes_test(is_admin)
def exams_session_detail(request, session_id):
    """Show detailed information about an exam session."""
    from exams.models import ExamSession
    
    session = get_object_or_404(ExamSession, pk=session_id)
    
    context = {
        'session': session,
    }
    
    return render(request, 'dashboard/exams/session_detail.html', context)


@login_required
@user_passes_test(is_admin)
def exams_schedules_list(request):
    """List all exam schedules."""
    from exams.models import ExamSchedule
    
    exam_schedules = ExamSchedule.objects.all().select_related('exam_session', 'course').order_by('exam_date', 'start_time')
    
    context = {
        'exam_schedules': exam_schedules,
    }
    
    return render(request, 'dashboard/exams/schedules_list.html', context)


@login_required
@user_passes_test(is_admin)
def exams_schedule_detail(request, schedule_id):
    """Show detailed information about an exam schedule."""
    from exams.models import ExamSchedule
    
    schedule = get_object_or_404(ExamSchedule, pk=schedule_id)
    
    context = {
        'schedule': schedule,
    }
    
    return render(request, 'dashboard/exams/schedule_detail.html', context)


@login_required
@user_passes_test(is_admin)
def exams_rooms_list(request):
    """List all exam rooms."""
    from exams.models import ExamRoom
    
    exam_rooms = ExamRoom.objects.all().order_by('building', 'floor', 'name')
    
    context = {
        'exam_rooms': exam_rooms,
    }
    
    return render(request, 'dashboard/exams/rooms_list.html', context)


@login_required
@user_passes_test(is_admin)
def exams_room_detail(request, room_id):
    """Show detailed information about an exam room."""
    from exams.models import ExamRoom
    
    room = get_object_or_404(ExamRoom, pk=room_id)
    
    context = {
        'room': room,
    }
    
    return render(request, 'dashboard/exams/room_detail.html', context)


@login_required
@user_passes_test(is_admin)
def exams_registrations_list(request):
    """List all exam registrations."""
    from exams.models import ExamRegistration
    
    exam_registrations = ExamRegistration.objects.all().select_related('student', 'exam_schedule__course').order_by('-registration_date')
    
    context = {
        'exam_registrations': exam_registrations,
    }
    
    return render(request, 'dashboard/exams/registrations_list.html', context)


@login_required
@user_passes_test(is_admin)
def exams_registration_detail(request, registration_id):
    """Show detailed information about an exam registration."""
    from exams.models import ExamRegistration
    
    registration = get_object_or_404(ExamRegistration, pk=registration_id)
    
    context = {
        'registration': registration,
    }
    
    return render(request, 'dashboard/exams/registration_detail.html', context)


@login_required
@user_passes_test(is_admin)
def exams_hall_tickets_list(request):
    """List all hall tickets."""
    from exams.models import HallTicket
    
    hall_tickets = HallTicket.objects.all().select_related('exam_registration__student', 'exam_registration__exam_schedule__course').order_by('-generated_date')
    
    context = {
        'hall_tickets': hall_tickets,
    }
    
    return render(request, 'dashboard/exams/hall_tickets_list.html', context)


@login_required
@user_passes_test(is_admin)
def exams_hall_ticket_detail(request, ticket_id):
    """Show detailed information about a hall ticket."""
    from exams.models import HallTicket
    
    ticket = get_object_or_404(HallTicket, pk=ticket_id)
    
    context = {
        'ticket': ticket,
    }
    
    return render(request, 'dashboard/exams/hall_ticket_detail.html', context)


@login_required
@user_passes_test(is_admin)
def exams_attendance_list(request):
    """List all exam attendance records."""
    from exams.models import ExamAttendance
    
    exam_attendance = ExamAttendance.objects.all().select_related('exam_registration__student', 'exam_registration__exam_schedule__course').order_by('-created_at')
    
    context = {
        'exam_attendance': exam_attendance,
    }
    
    return render(request, 'dashboard/exams/attendance_list.html', context)


@login_required
@user_passes_test(is_admin)
def exams_attendance_detail(request, attendance_id):
    """Show detailed information about an exam attendance record."""
    from exams.models import ExamAttendance
    
    attendance = get_object_or_404(ExamAttendance, pk=attendance_id)
    
    context = {
        'attendance': attendance,
    }
    
    return render(request, 'dashboard/exams/attendance_detail.html', context)


@login_required
@user_passes_test(is_admin)
def exams_results_list(request):
    """List all exam results."""
    from exams.models import ExamResult
    
    exam_results = ExamResult.objects.all().select_related('exam_registration__student', 'exam_registration__exam_schedule__course').order_by('-evaluated_at', '-created_at')
    
    context = {
        'exam_results': exam_results,
    }
    
    return render(request, 'dashboard/exams/results_list.html', context)


@login_required
@user_passes_test(is_admin)
def exams_result_detail(request, result_id):
    """Show detailed information about an exam result."""
    from exams.models import ExamResult
    
    result = get_object_or_404(ExamResult, pk=result_id)
    
    context = {
        'result': result,
    }
    
    return render(request, 'dashboard/exams/result_detail.html', context)


@login_required
@user_passes_test(is_admin)
def exams_dues_list(request):
    """List all student dues."""
    from exams.models import StudentDue
    
    student_dues = StudentDue.objects.all().select_related('student').order_by('-due_date')
    
    context = {
        'student_dues': student_dues,
    }
    
    return render(request, 'dashboard/exams/dues_list.html', context)


@login_required
@user_passes_test(is_admin)
def exams_due_detail(request, due_id):
    """Show detailed information about a student due."""
    from exams.models import StudentDue
    
    due = get_object_or_404(StudentDue, pk=due_id)
    
    context = {
        'due': due,
    }
    
    return render(request, 'dashboard/exams/due_detail.html', context)


@login_required
@user_passes_test(is_admin)
def exams_violations_list(request):
    """List all exam violations."""
    from exams.models import ExamViolation
    
    exam_violations = ExamViolation.objects.all().select_related('exam_registration__student', 'exam_registration__exam_schedule__course').order_by('-reported_at')
    
    context = {
        'exam_violations': exam_violations,
    }
    
    return render(request, 'dashboard/exams/violations_list.html', context)


@login_required
@user_passes_test(is_admin)
def exams_violation_detail(request, violation_id):
    """Show detailed information about an exam violation."""
    from exams.models import ExamViolation
    
    violation = get_object_or_404(ExamViolation, pk=violation_id)
    
    context = {
        'violation': violation,
    }
    
    return render(request, 'dashboard/exams/violation_detail.html', context)


@login_required
@user_passes_test(is_admin)
def exams_staff_assignments_list(request):
    """List all staff assignments."""
    from exams.models import ExamStaffAssignment
    
    staff_assignments = ExamStaffAssignment.objects.all().select_related('faculty', 'exam_schedule__course').order_by('-assigned_date')
    
    context = {
        'staff_assignments': staff_assignments,
    }
    
    return render(request, 'dashboard/exams/staff_assignments_list.html', context)


@login_required
@user_passes_test(is_admin)
def exams_staff_assignment_detail(request, assignment_id):
    """Show detailed information about a staff assignment."""
    from exams.models import ExamStaffAssignment
    
    assignment = get_object_or_404(ExamStaffAssignment, pk=assignment_id)
    
    context = {
        'assignment': assignment,
    }
    
    return render(request, 'dashboard/exams/staff_assignment_detail.html', context)


@login_required
@user_passes_test(is_admin)
def exams_room_allocations_list(request):
    """List all room allocations."""
    from exams.models import ExamRoomAllocation, ExamSchedule, ExamRoom
    from students.models import Student
    from departments.models import Department
    from academics.models import AcademicProgram
    
    room_allocations = ExamRoomAllocation.objects.all().select_related('exam_schedule__course', 'exam_room').order_by('-created_at')
    
    # Get data for the create modal
    students = Student.objects.filter(status='ACTIVE').order_by('roll_number')
    exam_schedules = ExamSchedule.objects.all().select_related('exam_session', 'course').order_by('-exam_date')
    exam_rooms = ExamRoom.objects.filter(is_active=True).order_by('building', 'name')
    
    # Get filter options
    departments = Department.objects.filter(is_active=True).order_by('name')
    academic_programs = AcademicProgram.objects.filter(is_active=True).order_by('name')
    
    # Get unique academic years and sections from students
    academic_years = Student.objects.values_list('academic_year', flat=True).distinct().exclude(academic_year__isnull=True).exclude(academic_year='').order_by('-academic_year')
    sections = Student.objects.values_list('section', flat=True).distinct().exclude(section__isnull=True).exclude(section='').order_by('section')
    
    context = {
        'room_allocations': room_allocations,
        'students': students,
        'exam_schedules': exam_schedules,
        'exam_rooms': exam_rooms,
        'departments': departments,
        'academic_programs': academic_programs,
        'academic_years': academic_years,
        'sections': sections,
    }
    
    return render(request, 'dashboard/exams/room_allocations_list.html', context)


@login_required
@user_passes_test(is_admin)
def exams_room_allocation_detail(request, allocation_id):
    """Show detailed information about a room allocation."""
    from exams.models import ExamRoomAllocation
    
    allocation = get_object_or_404(ExamRoomAllocation, pk=allocation_id)
    
    context = {
        'allocation': allocation,
    }
    
    return render(request, 'dashboard/exams/room_allocation_detail.html', context)

# =============================================================================
# FEE MANAGEMENT DASHBOARD VIEWS
# =============================================================================

@login_required
@user_passes_test(is_admin)
def fees_dashboard(request):
    """Main fee management dashboard with statistics and overview"""
    
    # Fee statistics
    total_fee_categories = FeeCategory.objects.filter(is_active=True).count()
    total_fee_structures = FeeStructure.objects.filter(is_active=True).count()
    total_student_fees = StudentFee.objects.count()
    total_payments = Payment.objects.filter(status='COMPLETED').count()
    
    # Financial statistics
    total_fees_due = StudentFee.objects.aggregate(total=models.Sum('amount_due'))['total'] or 0
    total_fees_paid = StudentFee.objects.aggregate(total=models.Sum('amount_paid'))['total'] or 0
    total_balance = total_fees_due - total_fees_paid
    
    # Status counts
    pending_fees = StudentFee.objects.filter(status='PENDING').count()
    paid_fees = StudentFee.objects.filter(status='PAID').count()
    partial_fees = StudentFee.objects.filter(status='PARTIAL').count()
    overdue_fees = StudentFee.objects.filter(due_date__lt=timezone.now().date(), status='PENDING').count()
    
    # Recent activities
    recent_payments = Payment.objects.filter(status='COMPLETED').select_related(
        'student_fee__student', 'collected_by'
    ).order_by('-payment_date')[:10]
    
    recent_student_fees = StudentFee.objects.select_related(
        'student', 'fee_structure_detail__fee_category'
    ).order_by('-created_at')[:10]
    
    # Payment method distribution
    payment_methods = Payment.objects.filter(status='COMPLETED').values('payment_method').annotate(
        count=Count('id'),
        total_amount=models.Sum('amount')
    ).order_by('-total_amount')
    
    # Fee category breakdown
    fee_category_breakdown = StudentFee.objects.values(
        'fee_structure_detail__fee_category__name'
    ).annotate(
        total_due=models.Sum('amount_due'),
        total_paid=models.Sum('amount_paid'),
        count=Count('id')
    ).order_by('-total_due')
    
    # Academic year breakdown
    academic_year_breakdown = StudentFee.objects.values('academic_year').annotate(
        total_due=models.Sum('amount_due'),
        total_paid=models.Sum('amount_paid'),
        count=Count('id')
    ).order_by('-academic_year')
    
    context = {
        'total_fee_categories': total_fee_categories,
        'total_fee_structures': total_fee_structures,
        'total_student_fees': total_student_fees,
        'total_payments': total_payments,
        'total_fees_due': total_fees_due,
        'total_fees_paid': total_fees_paid,
        'total_balance': total_balance,
        'pending_fees': pending_fees,
        'paid_fees': paid_fees,
        'partial_fees': partial_fees,
        'overdue_fees': overdue_fees,
        'recent_payments': recent_payments,
        'recent_student_fees': recent_student_fees,
        'payment_methods': payment_methods,
        'fee_category_breakdown': fee_category_breakdown,
        'academic_year_breakdown': academic_year_breakdown,
    }
    
    return render(request, 'dashboard/fees/dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def fees_categories_list(request):
    """List all fee categories"""
    categories = FeeCategory.objects.all().order_by('display_order', 'name')
    
    if request.method == 'POST':
        if 'create' in request.POST:
            name = request.POST.get('name')
            description = request.POST.get('description')
            display_order = request.POST.get('display_order', 0)
            is_active = 'is_active' in request.POST
            
            if name:
                FeeCategory.objects.create(
                    name=name,
                    description=description,
                    display_order=display_order,
                    is_active=is_active
                )
                return redirect('dashboard:fees_categories_list')
        
        elif 'update' in request.POST:
            category_id = request.POST.get('category_id')
            name = request.POST.get('name')
            description = request.POST.get('description')
            display_order = request.POST.get('display_order', 0)
            is_active = 'is_active' in request.POST
            
            if category_id and name:
                try:
                    category = FeeCategory.objects.get(id=category_id)
                    category.name = name
                    category.description = description
                    category.display_order = display_order
                    category.is_active = is_active
                    category.save()
                except FeeCategory.DoesNotExist:
                    pass
                return redirect('dashboard:fees_categories_list')
        
        elif 'delete' in request.POST:
            category_id = request.POST.get('category_id')
            if category_id:
                try:
                    FeeCategory.objects.get(id=category_id).delete()
                except FeeCategory.DoesNotExist:
                    pass
                return redirect('dashboard:fees_categories_list')
    
    context = {
        'categories': categories,
    }
    
    return render(request, 'dashboard/fees/categories_list.html', context)


@login_required
@user_passes_test(is_admin)
def fees_structures_list(request):
    """List all fee structures"""
    structures = FeeStructure.objects.all().select_related().prefetch_related('fee_details').order_by('-academic_year', 'grade_level')
    
    if request.method == 'POST':
        if 'create' in request.POST:
            name = request.POST.get('name')
            academic_year = request.POST.get('academic_year')
            grade_level = request.POST.get('grade_level')
            description = request.POST.get('description')
            is_active = 'is_active' in request.POST
            
            if name and academic_year and grade_level:
                FeeStructure.objects.create(
                    name=name,
                    academic_year=academic_year,
                    grade_level=grade_level,
                    description=description,
                    is_active=is_active
                )
                return redirect('dashboard:fees_structures_list')
        
        elif 'update' in request.POST:
            structure_id = request.POST.get('structure_id')
            name = request.POST.get('name')
            academic_year = request.POST.get('academic_year')
            grade_level = request.POST.get('grade_level')
            description = request.POST.get('description')
            is_active = 'is_active' in request.POST
            
            if structure_id and name and academic_year and grade_level:
                try:
                    structure = FeeStructure.objects.get(id=structure_id)
                    structure.name = name
                    structure.academic_year = academic_year
                    structure.grade_level = grade_level
                    structure.description = description
                    structure.is_active = is_active
                    structure.save()
                except FeeStructure.DoesNotExist:
                    pass
                return redirect('dashboard:fees_structures_list')
        
        elif 'delete' in request.POST:
            structure_id = request.POST.get('structure_id')
            if structure_id:
                try:
                    FeeStructure.objects.get(id=structure_id).delete()
                except FeeStructure.DoesNotExist:
                    pass
                return redirect('dashboard:fees_structures_list')
    
    context = {
        'structures': structures,
        'academic_year_choices': FeeStructure.ACADEMIC_YEAR_CHOICES,
        'grade_choices': FeeStructure.GRADE_CHOICES,
    }
    
    return render(request, 'dashboard/fees/structures_list.html', context)


@login_required
@user_passes_test(is_admin)
def fees_structure_detail(request, structure_id):
    """Show detailed information about a fee structure"""
    structure = get_object_or_404(FeeStructure, pk=structure_id)
    fee_details = structure.fee_details.all().select_related('fee_category').order_by('fee_category__display_order')
    
    if request.method == 'POST':
        if 'create_detail' in request.POST:
            fee_category_id = request.POST.get('fee_category_id')
            amount = request.POST.get('amount')
            frequency = request.POST.get('frequency')
            is_optional = 'is_optional' in request.POST
            due_date = request.POST.get('due_date') or None
            late_fee_amount = request.POST.get('late_fee_amount', 0)
            late_fee_percentage = request.POST.get('late_fee_percentage', 0)
            
            if fee_category_id and amount:
                try:
                    fee_category = FeeCategory.objects.get(id=fee_category_id)
                    FeeStructureDetail.objects.create(
                        fee_structure=structure,
                        fee_category=fee_category,
                        amount=amount,
                        frequency=frequency,
                        is_optional=is_optional,
                        due_date=due_date,
                        late_fee_amount=late_fee_amount,
                        late_fee_percentage=late_fee_percentage
                    )
                except FeeCategory.DoesNotExist:
                    pass
                return redirect('dashboard:fees_structure_detail', structure_id=structure_id)
        
        elif 'update_detail' in request.POST:
            detail_id = request.POST.get('detail_id')
            amount = request.POST.get('amount')
            frequency = request.POST.get('frequency')
            is_optional = 'is_optional' in request.POST
            due_date = request.POST.get('due_date') or None
            late_fee_amount = request.POST.get('late_fee_amount', 0)
            late_fee_percentage = request.POST.get('late_fee_percentage', 0)
            
            if detail_id and amount:
                try:
                    detail = FeeStructureDetail.objects.get(id=detail_id)
                    detail.amount = amount
                    detail.frequency = frequency
                    detail.is_optional = is_optional
                    detail.due_date = due_date
                    detail.late_fee_amount = late_fee_amount
                    detail.late_fee_percentage = late_fee_percentage
                    detail.save()
                except FeeStructureDetail.DoesNotExist:
                    pass
                return redirect('dashboard:fees_structure_detail', structure_id=structure_id)
        
        elif 'delete_detail' in request.POST:
            detail_id = request.POST.get('detail_id')
            if detail_id:
                try:
                    FeeStructureDetail.objects.get(id=detail_id).delete()
                except FeeStructureDetail.DoesNotExist:
                    pass
                return redirect('dashboard:fees_structure_detail', structure_id=structure_id)
    
    # Get available fee categories for creating new details
    available_categories = FeeCategory.objects.filter(is_active=True).exclude(
        id__in=fee_details.values_list('fee_category_id', flat=True)
    )
    
    context = {
        'structure': structure,
        'fee_details': fee_details,
        'available_categories': available_categories,
        'frequency_choices': FeeStructureDetail.FREQUENCY_CHOICES,
    }
    
    return render(request, 'dashboard/fees/structure_detail.html', context)


@login_required
@user_passes_test(is_admin)
def fees_student_fees_list(request):
    """List all student fees with filtering and search"""
    student_fees = StudentFee.objects.all().select_related(
        'student', 'fee_structure_detail__fee_category', 'fee_structure_detail__fee_structure'
    ).order_by('-created_at')
    
    # Filtering
    status_filter = request.GET.get('status')
    academic_year_filter = request.GET.get('academic_year')
    grade_filter = request.GET.get('grade_level')
    category_filter = request.GET.get('fee_category')
    search_query = request.GET.get('search')
    
    if status_filter:
        student_fees = student_fees.filter(status=status_filter)
    
    if academic_year_filter:
        student_fees = student_fees.filter(academic_year=academic_year_filter)
    
    if grade_filter:
        student_fees = student_fees.filter(fee_structure_detail__fee_structure__grade_level=grade_filter)
    
    if category_filter:
        student_fees = student_fees.filter(fee_structure_detail__fee_category_id=category_filter)
    
    if search_query:
        student_fees = student_fees.filter(
            Q(student__roll_number__icontains=search_query) |
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query)
        )
    
    # Get filter options
    status_choices = StudentFee.STATUS_CHOICES
    academic_years = StudentFee.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')
    grade_levels = FeeStructure.objects.values_list('grade_level', flat=True).distinct().order_by('grade_level')
    fee_categories = FeeCategory.objects.filter(is_active=True).order_by('name')
    
    context = {
        'student_fees': student_fees,
        'status_choices': status_choices,
        'academic_years': academic_years,
        'grade_levels': grade_levels,
        'fee_categories': fee_categories,
        'current_filters': {
            'status': status_filter,
            'academic_year': academic_year_filter,
            'grade_level': grade_filter,
            'fee_category': category_filter,
            'search': search_query,
        }
    }
    
    return render(request, 'dashboard/fees/student_fees_list.html', context)


@login_required
@user_passes_test(is_admin)
def fees_student_fee_detail(request, student_fee_id):
    """Show detailed information about a student fee"""
    student_fee = get_object_or_404(StudentFee, pk=student_fee_id)
    payments = student_fee.payments.all().order_by('-payment_date')
    waivers = student_fee.waivers.filter(is_active=True).order_by('-created_at')
    discounts = student_fee.discounts.filter(is_active=True).order_by('-created_at')
    receipts = student_fee.receipts.all().order_by('-generated_date')
    
    context = {
        'student_fee': student_fee,
        'payments': payments,
        'waivers': waivers,
        'discounts': discounts,
        'receipts': receipts,
    }
    
    return render(request, 'dashboard/fees/student_fee_detail.html', context)


@login_required
@user_passes_test(is_admin)
def fees_payments_list(request):
    """List all payments with filtering and search"""
    payments = Payment.objects.all().select_related(
        'student_fee__student', 'student_fee__fee_structure_detail__fee_category', 'collected_by'
    ).order_by('-payment_date')
    
    # Filtering
    status_filter = request.GET.get('status')
    method_filter = request.GET.get('payment_method')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search_query = request.GET.get('search')
    
    if status_filter:
        payments = payments.filter(status=status_filter)
    
    if method_filter:
        payments = payments.filter(payment_method=method_filter)
    
    if date_from:
        payments = payments.filter(payment_date__date__gte=date_from)
    
    if date_to:
        payments = payments.filter(payment_date__date__lte=date_to)
    
    if search_query:
        payments = payments.filter(
            Q(receipt_number__icontains=search_query) |
            Q(transaction_id__icontains=search_query) |
            Q(student_fee__student__roll_number__icontains=search_query) |
            Q(student_fee__student__first_name__icontains=search_query)
        )
    
    # Get filter options
    status_choices = Payment.PAYMENT_STATUS_CHOICES
    method_choices = Payment.PAYMENT_METHOD_CHOICES
    
    context = {
        'payments': payments,
        'status_choices': status_choices,
        'method_choices': method_choices,
        'current_filters': {
            'status': status_filter,
            'payment_method': method_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search_query,
        }
    }
    
    return render(request, 'dashboard/fees/payments_list.html', context)


@login_required
@user_passes_test(is_admin)
def fees_payment_detail(request, payment_id):
    """Show detailed information about a payment"""
    payment = get_object_or_404(Payment, pk=payment_id)
    receipts = payment.receipts.all().order_by('-generated_date')
    
    context = {
        'payment': payment,
        'receipts': receipts,
    }
    
    return render(request, 'dashboard/fees/payment_detail.html', context)


@login_required
@user_passes_test(is_admin)
def fees_waivers_list(request):
    """List all fee waivers"""
    waivers = FeeWaiver.objects.all().select_related(
        'student_fee__student', 'approved_by'
    ).order_by('-created_at')
    
    if request.method == 'POST':
        if 'create' in request.POST:
            student_fee_id = request.POST.get('student_fee_id')
            waiver_type = request.POST.get('waiver_type')
            amount = request.POST.get('amount')
            percentage = request.POST.get('percentage')
            reason = request.POST.get('reason')
            is_active = 'is_active' in request.POST
            
            if student_fee_id and waiver_type and amount and reason:
                try:
                    student_fee = StudentFee.objects.get(id=student_fee_id)
                    FeeWaiver.objects.create(
                        student_fee=student_fee,
                        waiver_type=waiver_type,
                        amount=amount,
                        percentage=percentage,
                        reason=reason,
                        is_active=is_active,
                        approved_by=request.user,
                        approved_date=timezone.now()
                    )
                except StudentFee.DoesNotExist:
                    pass
                return redirect('dashboard:fees_waivers_list')
        
        elif 'update' in request.POST:
            waiver_id = request.POST.get('waiver_id')
            waiver_type = request.POST.get('waiver_type')
            amount = request.POST.get('amount')
            percentage = request.POST.get('percentage')
            reason = request.POST.get('reason')
            is_active = 'is_active' in request.POST
            
            if waiver_id and waiver_type and amount and reason:
                try:
                    waiver = FeeWaiver.objects.get(id=waiver_id)
                    waiver.waiver_type = waiver_type
                    waiver.amount = amount
                    waiver.percentage = percentage
                    waiver.reason = reason
                    waiver.is_active = is_active
                    waiver.save()
                except FeeWaiver.DoesNotExist:
                    pass
                return redirect('dashboard:fees_waivers_list')
        
        elif 'delete' in request.POST:
            waiver_id = request.POST.get('waiver_id')
            if waiver_id:
                try:
                    FeeWaiver.objects.get(id=waiver_id).delete()
                except FeeWaiver.DoesNotExist:
                    pass
                return redirect('dashboard:fees_waivers_list')
    
    # Get available student fees for creating new waivers
    available_student_fees = StudentFee.objects.filter(status='PENDING').select_related('student')
    
    context = {
        'waivers': waivers,
        'available_student_fees': available_student_fees,
        'waiver_type_choices': FeeWaiver.WAIVER_TYPE_CHOICES,
    }
    
    return render(request, 'dashboard/fees/waivers_list.html', context)


@login_required
@user_passes_test(is_admin)
def fees_discounts_list(request):
    """List all fee discounts"""
    discounts = FeeDiscount.objects.all().select_related(
        'student_fee__student'
    ).order_by('-created_at')
    
    if request.method == 'POST':
        if 'create' in request.POST:
            student_fee_id = request.POST.get('student_fee_id')
            discount_type = request.POST.get('discount_type')
            amount = request.POST.get('amount')
            percentage = request.POST.get('percentage')
            reason = request.POST.get('reason')
            valid_until = request.POST.get('valid_until') or None
            is_active = 'is_active' in request.POST
            
            if student_fee_id and discount_type and amount and reason:
                try:
                    student_fee = StudentFee.objects.get(id=student_fee_id)
                    FeeDiscount.objects.create(
                        student_fee=student_fee,
                        discount_type=discount_type,
                        amount=amount,
                        percentage=percentage,
                        reason=reason,
                        valid_until=valid_until,
                        is_active=is_active
                    )
                except StudentFee.DoesNotExist:
                    pass
                return redirect('dashboard:fees_discounts_list')
        
        elif 'update' in request.POST:
            discount_id = request.POST.get('discount_id')
            discount_type = request.POST.get('discount_type')
            amount = request.POST.get('amount')
            percentage = request.POST.get('percentage')
            reason = request.POST.get('reason')
            valid_until = request.POST.get('valid_until') or None
            is_active = 'is_active' in request.POST
            
            if discount_id and discount_type and amount and reason:
                try:
                    discount = FeeDiscount.objects.get(id=discount_id)
                    discount.discount_type = discount_type
                    discount.amount = amount
                    discount.percentage = percentage
                    discount.reason = reason
                    discount.valid_until = valid_until
                    discount.is_active = is_active
                    discount.save()
                except FeeDiscount.DoesNotExist:
                    pass
                return redirect('dashboard:fees_discounts_list')
        
        elif 'delete' in request.POST:
            discount_id = request.POST.get('discount_id')
            if discount_id:
                try:
                    FeeDiscount.objects.get(id=discount_id).delete()
                except FeeDiscount.DoesNotExist:
                    pass
                return redirect('dashboard:fees_discounts_list')
    
    # Get available student fees for creating new discounts
    available_student_fees = StudentFee.objects.filter(status='PENDING').select_related('student')
    
    context = {
        'discounts': discounts,
        'available_student_fees': available_student_fees,
        'discount_type_choices': FeeDiscount.DISCOUNT_TYPE_CHOICES,
    }
    
    return render(request, 'dashboard/fees/discounts_list.html', context)


@login_required
@user_passes_test(is_admin)
def fees_receipts_list(request):
    """List all fee receipts"""
    receipts = FeeReceipt.objects.all().select_related(
        'student_fee__student', 'payment', 'generated_by'
    ).order_by('-generated_date')
    
    # Filtering
    printed_filter = request.GET.get('printed')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search_query = request.GET.get('search')
    
    if printed_filter is not None:
        receipts = receipts.filter(is_printed=printed_filter == 'true')
    
    if date_from:
        receipts = receipts.filter(generated_date__date__gte=date_from)
    
    if date_to:
        receipts = receipts.filter(generated_date__date__lte=date_to)
    
    if search_query:
        receipts = receipts.filter(
            Q(receipt_number__icontains=search_query) |
            Q(student_fee__student__roll_number__icontains=search_query) |
            Q(student_fee__student__first_name__icontains=search_query)
        )
    
    context = {
        'receipts': receipts,
        'current_filters': {
            'printed': printed_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search_query,
        }
    }
    
    return render(request, 'dashboard/fees/receipts_list.html', context)


@login_required
@user_passes_test(is_admin)
def fees_receipt_detail(request, receipt_id):
    """Show detailed information about a receipt"""
    receipt = get_object_or_404(FeeReceipt, pk=receipt_id)
    
    context = {
        'receipt': receipt,
    }
    
    return render(request, 'dashboard/fees/receipt_detail.html', context)


@login_required
@user_passes_test(is_admin)
def fees_reports(request):
    """Fee management reports and analytics"""
    
    # Date range filtering
    date_from = request.GET.get('date_from', (timezone.now() - timezone.timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))
    
    # Convert to date objects
    try:
        from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
        to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
    except ValueError:
        from_date = (timezone.now() - timezone.timedelta(days=30)).date()
        to_date = timezone.now().date()
    
    # Payment reports
    payments_in_range = Payment.objects.filter(
        payment_date__date__range=[from_date, to_date],
        status='COMPLETED'
    )
    
    total_payments_amount = payments_in_range.aggregate(total=models.Sum('amount'))['total'] or 0
    total_payments_count = payments_in_range.count()
    
    # Payment method breakdown
    payment_method_breakdown = payments_in_range.values('payment_method').annotate(
        count=Count('id'),
        total_amount=models.Sum('amount')
    ).order_by('-total_amount')
    
    # Daily payment trends
    daily_payments = payments_in_range.values('payment_date__date').annotate(
        count=Count('id'),
        total_amount=models.Sum('amount')
    ).order_by('payment_date__date')
    
    # Fee category performance
    fee_category_performance = StudentFee.objects.filter(
        created_at__date__range=[from_date, to_date]
    ).values('fee_structure_detail__fee_category__name').annotate(
        total_due=models.Sum('amount_due'),
        total_paid=models.Sum('amount_paid'),
        count=Count('id')
    ).order_by('-total_due')
    
    # Academic year performance
    academic_year_performance = StudentFee.objects.filter(
        created_at__date__range=[from_date, to_date]
    ).values('academic_year').annotate(
        total_due=models.Sum('amount_due'),
        total_paid=models.Sum('amount_paid'),
        count=Count('id')
    ).order_by('-academic_year')
    
    # Overdue fees report
    overdue_fees = StudentFee.objects.filter(
        due_date__lt=timezone.now().date(),
        status='PENDING'
    ).select_related('student', 'fee_structure_detail__fee_category')
    
    overdue_amount = overdue_fees.aggregate(total=models.Sum('amount_due'))['total'] or 0
    
    # Waiver and discount reports
    waivers_in_range = FeeWaiver.objects.filter(
        created_at__date__range=[from_date, to_date],
        is_active=True
    )
    
    discounts_in_range = FeeDiscount.objects.filter(
        created_at__date__range=[from_date, to_date],
        is_active=True
    )
    
    total_waiver_amount = waivers_in_range.aggregate(total=models.Sum('amount'))['total'] or 0
    total_discount_amount = discounts_in_range.aggregate(total=models.Sum('amount'))['total'] or 0
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'total_payments_amount': total_payments_amount,
        'total_payments_count': total_payments_count,
        'payment_method_breakdown': payment_method_breakdown,
        'daily_payments': daily_payments,
        'fee_category_performance': fee_category_performance,
        'academic_year_performance': academic_year_performance,
        'overdue_fees': overdue_fees,
        'overdue_amount': overdue_amount,
        'total_waiver_amount': total_waiver_amount,
        'total_discount_amount': total_discount_amount,
    }
    
    return render(request, 'dashboard/fees/reports.html', context)


@login_required
@user_passes_test(is_admin)
def fees_api_endpoints(request):
    """Show available fee management API endpoints"""
    
    context = {
        'app_name': 'Fee Management',
        'base_url': '/api/v1/fees/api/',
        'endpoints': [
            {
                'name': 'Fee Categories',
                'url': 'categories/',
                'methods': ['GET', 'POST'],
                'description': 'Manage fee categories (Tuition, Library, Sports, etc.)'
            },
            {
                'name': 'Fee Structures',
                'url': 'structures/',
                'methods': ['GET', 'POST'],
                'description': 'Manage fee structures for different academic years and grades'
            },
            {
                'name': 'Fee Structure Details',
                'url': 'structure-details/',
                'methods': ['GET', 'POST'],
                'description': 'Manage individual fee items within structures'
            },
            {
                'name': 'Student Fees',
                'url': 'student-fees/',
                'methods': ['GET', 'POST'],
                'description': 'Manage individual student fee records'
            },
            {
                'name': 'Payments',
                'url': 'payments/',
                'methods': ['GET', 'POST'],
                'description': 'Manage fee payments and transactions'
            },
            {
                'name': 'Fee Waivers',
                'url': 'waivers/',
                'methods': ['GET', 'POST'],
                'description': 'Manage fee waivers and scholarships'
            },
            {
                'name': 'Fee Discounts',
                'url': 'discounts/',
                'methods': ['GET', 'POST'],
                'description': 'Manage fee discounts and reductions'
            },
            {
                'name': 'Fee Receipts',
                'url': 'receipts/',
                'methods': ['GET', 'POST'],
                'description': 'Manage fee receipts and documentation'
            },
        ]
    }
    
    return render(request, 'dashboard/fees/api_endpoints.html', context)

# ------------------------------
# Open Requests (UI)
# ------------------------------
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
try:
    from open_requests.models import OpenRequest, RequestTarget
except Exception:
    OpenRequest = RequestTarget = None


@login_required
def open_requests_list(request):
    if OpenRequest is None:
        return redirect('dashboard:home')
    queryset = OpenRequest.objects.select_related('created_by', 'assignee').all()
    paginator = Paginator(queryset, 20)
    page = request.GET.get('page')
    items = paginator.get_page(page)
    return render(request, 'dashboard/open_requests/list.html', {'items': items})


@login_required
def open_request_create(request):
    if OpenRequest is None:
        return redirect('dashboard:home')
    if request.method == 'POST':
        OpenRequest.objects.create(
            title=request.POST.get('title',''),
            description=request.POST.get('description',''),
            created_by=request.user,
        )
        return redirect('dashboard:open_requests_list')
    return render(request, 'dashboard/open_requests/create.html')


@login_required
def open_request_detail(request, request_id):
    if OpenRequest is None:
        return redirect('dashboard:home')
    item = get_object_or_404(OpenRequest, id=request_id)
    return render(request, 'dashboard/open_requests/detail.html', {'item': item})


def custom_logout(request):
    """Custom logout view that handles both GET and POST requests"""
    if request.user.is_authenticated:
        logout(request)
        messages.success(request, 'You have been successfully logged out.')
    return redirect('dashboard:login')


# ==================== ASSIGNMENTS DASHBOARD VIEWS ====================

@login_required
def assignments_dashboard(request):
    """Main assignments dashboard"""
    user = request.user
    
    # Get assignments based on user role
    if hasattr(user, 'faculty_profile'):
        # Faculty dashboard
        faculty = user.faculty_profile
        assignments = Assignment.objects.filter(faculty=faculty).order_by('-created_at')
        
        stats = {
            'total_assignments': assignments.count(),
            'published_assignments': assignments.filter(status='PUBLISHED').count(),
            'draft_assignments': assignments.filter(status='DRAFT').count(),
            'overdue_assignments': assignments.filter(
                status='PUBLISHED',
                due_date__lt=timezone.now()
            ).count(),
            'total_submissions': AssignmentSubmission.objects.filter(
                assignment__faculty=faculty
            ).count(),
            'graded_submissions': AssignmentSubmission.objects.filter(
                assignment__faculty=faculty,
                grade__isnull=False
            ).count(),
            'pending_grades': AssignmentSubmission.objects.filter(
                assignment__faculty=faculty,
                grade__isnull=True
            ).count(),
        }
        
        recent_assignments = assignments[:5]
        recent_submissions = AssignmentSubmission.objects.filter(
            assignment__faculty=faculty
        ).order_by('-submission_date')[:5]
        
    elif hasattr(user, 'student_profile'):
        # Student dashboard
        student = user.student_profile
        assignments = Assignment.objects.filter(
            Q(assigned_to_students=student) | 
            Q(assigned_to_grades__students=student)
        ).distinct().order_by('-created_at')
        
        stats = {
            'total_assignments': assignments.count(),
            'submitted_assignments': AssignmentSubmission.objects.filter(student=student).count(),
            'pending_assignments': assignments.exclude(
                submissions__student=student
            ).count(),
            'late_submissions': AssignmentSubmission.objects.filter(
                student=student, is_late=True
            ).count(),
        }
        
        recent_assignments = assignments[:5]
        my_submissions = AssignmentSubmission.objects.filter(
            student=student
        ).order_by('-submission_date')[:5]
        
    else:
        # Admin dashboard
        assignments = Assignment.objects.all().order_by('-created_at')
        
        stats = {
            'total_assignments': assignments.count(),
            'published_assignments': assignments.filter(status='PUBLISHED').count(),
            'draft_assignments': assignments.filter(status='DRAFT').count(),
            'overdue_assignments': assignments.filter(
                status='PUBLISHED',
                due_date__lt=timezone.now()
            ).count(),
            'total_submissions': AssignmentSubmission.objects.count(),
            'graded_submissions': AssignmentSubmission.objects.filter(
                grade__isnull=False
            ).count(),
            'pending_grades': AssignmentSubmission.objects.filter(
                grade__isnull=True
            ).count(),
        }
        
        recent_assignments = assignments[:5]
        recent_submissions = AssignmentSubmission.objects.all().order_by('-submission_date')[:5]
    
    context = {
        'stats': stats,
        'recent_assignments': recent_assignments,
        'categories': AssignmentCategory.objects.filter(is_active=True),
        'user_role': 'faculty' if hasattr(user, 'faculty_profile') else 'student' if hasattr(user, 'student_profile') else 'admin',
    }
    
    if hasattr(user, 'faculty_profile'):
        context['recent_submissions'] = recent_submissions
    elif hasattr(user, 'student_profile'):
        context['my_submissions'] = my_submissions
    
    return render(request, 'dashboard/assignments/dashboard.html', context)


@login_required
def assignments_list(request):
    """List all assignments with filtering"""
    user = request.user
    
    # Get assignments based on user role
    if hasattr(user, 'faculty_profile'):
        assignments = Assignment.objects.filter(faculty=user.faculty_profile)
    elif hasattr(user, 'student_profile'):
        student = user.student_profile
        assignments = Assignment.objects.filter(
            Q(assigned_to_students=student) | 
            Q(assigned_to_grades__students=student)
        ).distinct()
    else:
        assignments = Assignment.objects.all()
    
    # Apply filters
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    category_filter = request.GET.get('category', '')
    is_overdue = request.GET.get('overdue', '')
    
    if search:
        assignments = assignments.filter(
            Q(title__icontains=search) | 
            Q(description__icontains=search)
        )
    
    if status_filter:
        assignments = assignments.filter(status=status_filter)
    
    if category_filter:
        assignments = assignments.filter(category_id=category_filter)
    
    if is_overdue == 'true':
        assignments = assignments.filter(
            status='PUBLISHED',
            due_date__lt=timezone.now()
        )
    
    assignments = assignments.order_by('-created_at')
    
    context = {
        'assignments': assignments,
        'categories': AssignmentCategory.objects.filter(is_active=True),
        'status_choices': Assignment.STATUS_CHOICES,
        'filters': {
            'search': search,
            'status': status_filter,
            'category': category_filter,
            'overdue': is_overdue,
        }
    }
    
    return render(request, 'dashboard/assignments/list.html', context)


@login_required
def assignment_detail(request, assignment_id):
    """Assignment detail view"""
    assignment = get_object_or_404(Assignment, id=assignment_id)
    user = request.user
    
    # Check permissions
    if hasattr(user, 'student_profile'):
        student = user.student_profile
        if not (assignment.assigned_to_students.filter(id=student.id).exists() or 
                assignment.assigned_to_grades.filter(students=student).exists()):
            messages.error(request, 'You do not have permission to view this assignment.')
            return redirect('dashboard:assignments_list')
    elif hasattr(user, 'faculty_profile'):
        if assignment.faculty != user.faculty_profile:
            messages.error(request, 'You do not have permission to view this assignment.')
            return redirect('dashboard:assignments_list')
    
    # Get submissions if user is faculty or admin
    submissions = None
    if hasattr(user, 'faculty_profile') or user.is_staff:
        submissions = AssignmentSubmission.objects.filter(
            assignment=assignment
        ).order_by('-submission_date')
    
    # Get user's submission if student
    my_submission = None
    if hasattr(user, 'student_profile'):
        my_submission = AssignmentSubmission.objects.filter(
            assignment=assignment,
            student=user.student_profile
        ).first()
    
    # Get comments
    comments = AssignmentComment.objects.filter(
        assignment=assignment,
        parent_comment__isnull=True
    ).order_by('created_at')
    
    context = {
        'assignment': assignment,
        'submissions': submissions,
        'my_submission': my_submission,
        'comments': comments,
        'can_edit': hasattr(user, 'faculty_profile') and assignment.faculty == user.faculty_profile,
        'can_submit': hasattr(user, 'student_profile') and not my_submission,
        'can_grade': hasattr(user, 'faculty_profile') and assignment.faculty == user.faculty_profile,
    }
    
    return render(request, 'dashboard/assignments/detail.html', context)


@login_required
def assignment_create(request):
    """Create new assignment"""
    if not hasattr(request.user, 'faculty_profile'):
        messages.error(request, 'Only faculty can create assignments.')
        return redirect('dashboard:assignments_list')
    
    if request.method == 'POST':
        form = AssignmentForm(request.POST)
        if form.is_valid():
            try:
                assignment = form.save(commit=False)
                assignment.faculty = request.user.faculty_profile
                assignment.status = 'DRAFT'
                assignment.save()
                form.save_m2m()  # Save many-to-many relationships
                
                messages.success(request, 'Assignment created successfully.')
                return redirect('dashboard:assignment_detail', assignment_id=assignment.id)
                
            except Exception as e:
                messages.error(request, f'Error creating assignment: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AssignmentForm()
    
    context = {
        'form': form,
        'categories': AssignmentCategory.objects.filter(is_active=True),
        'programs': apps.get_model('academics', 'AcademicProgram').objects.filter(is_active=True) if apps.is_installed('academics') else [],
        'departments': apps.get_model('departments', 'Department').objects.filter(is_active=True) if apps.is_installed('departments') else [],
        'courses': apps.get_model('academics', 'Course').objects.filter(status='ACTIVE') if apps.is_installed('academics') else [],
        'course_sections': apps.get_model('academics', 'CourseSection').objects.filter(is_active=True) if apps.is_installed('academics') else [],
        'students': Student.objects.all()[:100],  # Limit for performance
    }
    
    return render(request, 'dashboard/assignments/create.html', context)


@login_required
def assignment_edit(request, assignment_id):
    """Edit assignment"""
    assignment = get_object_or_404(Assignment, id=assignment_id)
    
    if not hasattr(request.user, 'faculty_profile') or assignment.faculty != request.user.faculty_profile:
        messages.error(request, 'You do not have permission to edit this assignment.')
        return redirect('dashboard:assignments_list')
    
    if request.method == 'POST':
        # Handle form submission
        assignment.title = request.POST.get('title', '').strip()
        assignment.description = request.POST.get('description', '').strip()
        assignment.instructions = request.POST.get('instructions', '').strip()
        assignment.max_marks = float(request.POST.get('max_marks', '0'))
        assignment.due_date = timezone.datetime.fromisoformat(
            request.POST.get('due_date', '').replace('Z', '+00:00')
        )
        assignment.category_id = request.POST.get('category') or None
        assignment.is_group_assignment = request.POST.get('is_group_assignment') == 'on'
        assignment.max_group_size = int(request.POST.get('max_group_size', '1'))
        assignment.late_submission_allowed = request.POST.get('late_submission_allowed') == 'on'
        
        # Handle assigned programs, departments, course sections, and students
        assigned_programs = request.POST.getlist('assigned_programs')
        assigned_departments = request.POST.getlist('assigned_departments')
        assigned_course_sections = request.POST.getlist('assigned_course_sections')
        assigned_students = request.POST.getlist('assigned_students')
        
        assignment.save()
        
        if assigned_programs:
            assignment.assigned_to_programs.set(assigned_programs)
        if assigned_departments:
            assignment.assigned_to_departments.set(assigned_departments)
        if assigned_course_sections:
            assignment.assigned_to_course_sections.set(assigned_course_sections)
        if assigned_students:
            assignment.assigned_to_students.set(assigned_students)
        
        messages.success(request, 'Assignment updated successfully.')
        return redirect('dashboard:assignment_detail', assignment_id=assignment.id)
    
    context = {
        'assignment': assignment,
        'categories': AssignmentCategory.objects.filter(is_active=True),
        'programs': apps.get_model('academics', 'AcademicProgram').objects.filter(is_active=True) if apps.is_installed('academics') else [],
        'departments': apps.get_model('academics', 'Department').objects.filter(is_active=True) if apps.is_installed('academics') else [],
        'course_sections': apps.get_model('academics', 'CourseSection').objects.filter(is_active=True) if apps.is_installed('academics') else [],
        'students': Student.objects.all()[:100],
    }
    
    return render(request, 'dashboard/assignments/edit.html', context)


@login_required
def assignment_submit(request, assignment_id):
    """Submit assignment"""
    assignment = get_object_or_404(Assignment, id=assignment_id)
    
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'Only students can submit assignments.')
        return redirect('dashboard:assignments_list')
    
    student = request.user.student_profile
    
    # Check if assignment is assigned to student
    if not (assignment.assigned_to_students.filter(id=student.id).exists() or 
            assignment.assigned_to_grades.filter(students=student).exists()):
        messages.error(request, 'This assignment is not assigned to you.')
        return redirect('dashboard:assignments_list')
    
    # Check if already submitted
    existing_submission = AssignmentSubmission.objects.filter(
        assignment=assignment,
        student=student
    ).first()
    
    if existing_submission:
        messages.error(request, 'You have already submitted this assignment.')
        return redirect('dashboard:assignment_detail', assignment_id=assignment.id)
    
    if request.method == 'POST':
        content = request.POST.get('content', '').strip()
        notes = request.POST.get('notes', '').strip()
        
        if not content:
            messages.error(request, 'Please provide your submission content.')
        else:
            try:
                submission = AssignmentSubmission.objects.create(
                    assignment=assignment,
                    student=student,
                    content=content,
                    notes=notes,
                    status='SUBMITTED'
                )
                
                messages.success(request, 'Assignment submitted successfully.')
                return redirect('dashboard:assignment_detail', assignment_id=assignment.id)
                
            except Exception as e:
                messages.error(request, f'Error submitting assignment: {str(e)}')
    
    return render(request, 'dashboard/assignments/submit.html', {'assignment': assignment})


@login_required
def assignment_grade(request, submission_id):
    """Grade assignment submission"""
    submission = get_object_or_404(AssignmentSubmission, id=submission_id)
    
    if not hasattr(request.user, 'faculty_profile') or submission.assignment.faculty != request.user.faculty_profile:
        messages.error(request, 'You do not have permission to grade this submission.')
        return redirect('dashboard:assignments_list')
    
    if request.method == 'POST':
        marks_obtained = request.POST.get('marks_obtained', '')
        grade_letter = request.POST.get('grade_letter', '')
        feedback = request.POST.get('feedback', '').strip()
        
        if not marks_obtained:
            messages.error(request, 'Please provide marks obtained.')
        else:
            try:
                # Create or update grade
                grade, created = AssignmentGrade.objects.get_or_create(
                    submission=submission,
                    defaults={
                        'marks_obtained': float(marks_obtained),
                        'grade_letter': grade_letter if grade_letter else None,
                        'feedback': feedback,
                        'graded_by': request.user
                    }
                )
                
                if not created:
                    grade.marks_obtained = float(marks_obtained)
                    grade.grade_letter = grade_letter if grade_letter else None
                    grade.feedback = feedback
                    grade.graded_by = request.user
                    grade.save()
                
                # Update submission
                submission.grade = grade
                submission.graded_by = request.user
                submission.graded_at = timezone.now()
                submission.feedback = feedback
                submission.save()
                
                messages.success(request, 'Assignment graded successfully.')
                return redirect('dashboard:assignment_detail', assignment_id=submission.assignment.id)
                
            except Exception as e:
                messages.error(request, f'Error grading assignment: {str(e)}')
    
    context = {
        'submission': submission,
        'grade_choices': AssignmentGrade.GRADE_LETTER_CHOICES,
    }
    
    return render(request, 'dashboard/assignments/grade.html', context)


@login_required
def assignment_categories(request):
    """Manage assignment categories"""
    if not request.user.is_staff:
        messages.error(request, 'Only administrators can manage categories.')
        return redirect('dashboard:assignments_dashboard')
    
    categories = AssignmentCategory.objects.all().order_by('name')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            color_code = request.POST.get('color_code', '#007bff')
            
            if name:
                try:
                    AssignmentCategory.objects.create(
                        name=name,
                        description=description,
                        color_code=color_code
                    )
                    messages.success(request, 'Category created successfully.')
                except Exception as e:
                    messages.error(request, f'Error creating category: {str(e)}')
            else:
                messages.error(request, 'Category name is required.')
        
        elif action == 'update':
            category_id = request.POST.get('category_id')
            category = get_object_or_404(AssignmentCategory, id=category_id)
            
            category.name = request.POST.get('name', '').strip()
            category.description = request.POST.get('description', '').strip()
            category.color_code = request.POST.get('color_code', '#007bff')
            category.is_active = request.POST.get('is_active') == 'on'
            
            try:
                category.save()
                messages.success(request, 'Category updated successfully.')
            except Exception as e:
                messages.error(request, f'Error updating category: {str(e)}')
        
        elif action == 'delete':
            category_id = request.POST.get('category_id')
            category = get_object_or_404(AssignmentCategory, id=category_id)
            
            try:
                category.delete()
                messages.success(request, 'Category deleted successfully.')
            except Exception as e:
                messages.error(request, f'Error deleting category: {str(e)}')
        
        return redirect('dashboard:assignment_categories')
    
    return render(request, 'dashboard/assignments/categories.html', {'categories': categories})


@login_required
def assignment_templates(request):
    """Manage assignment templates"""
    if not hasattr(request.user, 'faculty_profile') and not request.user.is_staff:
        messages.error(request, 'Only faculty can manage templates.')
        return redirect('dashboard:assignments_dashboard')
    
    # Get templates accessible to user
    if request.user.is_staff:
        templates = AssignmentTemplate.objects.all()
    else:
        templates = AssignmentTemplate.objects.filter(
            Q(is_public=True) | Q(created_by=request.user)
        )
    
    templates = templates.order_by('-created_at')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            instructions = request.POST.get('instructions', '').strip()
            max_marks = request.POST.get('max_marks', '')
            category_id = request.POST.get('category', '')
            is_group_assignment = request.POST.get('is_group_assignment') == 'on'
            max_group_size = request.POST.get('max_group_size', '1')
            is_public = request.POST.get('is_public') == 'on'
            
            if name and max_marks:
                try:
                    AssignmentTemplate.objects.create(
                        name=name,
                        description=description,
                        instructions=instructions,
                        max_marks=float(max_marks),
                        category_id=category_id if category_id else None,
                        is_group_assignment=is_group_assignment,
                        max_group_size=int(max_group_size) if max_group_size else 1,
                        is_public=is_public,
                        created_by=request.user
                    )
                    messages.success(request, 'Template created successfully.')
                except Exception as e:
                    messages.error(request, f'Error creating template: {str(e)}')
            else:
                messages.error(request, 'Template name and max marks are required.')
        
        return redirect('dashboard:assignment_templates')
    
    context = {
        'templates': templates,
        'categories': AssignmentCategory.objects.filter(is_active=True),
    }
    
    return render(request, 'dashboard/assignments/templates.html', context)


@login_required
def assignment_statistics(request):
    """Assignment statistics and analytics"""
    user = request.user
    
    if hasattr(user, 'faculty_profile'):
        # Faculty statistics
        faculty = user.faculty_profile
        assignments = Assignment.objects.filter(faculty=faculty)
        submissions = AssignmentSubmission.objects.filter(assignment__faculty=faculty)
        
        stats = {
            'total_assignments': assignments.count(),
            'published_assignments': assignments.filter(status='PUBLISHED').count(),
            'draft_assignments': assignments.filter(status='DRAFT').count(),
            'overdue_assignments': assignments.filter(
                status='PUBLISHED',
                due_date__lt=timezone.now()
            ).count(),
            'total_submissions': submissions.count(),
            'graded_submissions': submissions.filter(grade__isnull=False).count(),
            'pending_grades': submissions.filter(grade__isnull=True).count(),
            'average_grade': submissions.filter(grade__isnull=False).aggregate(
                avg=Avg('grade__marks_obtained')
            )['avg'] or 0,
        }
        
    elif hasattr(user, 'student_profile'):
        # Student statistics
        student = user.student_profile
        submissions = AssignmentSubmission.objects.filter(student=student)
        
        stats = {
            'total_assignments': Assignment.objects.filter(
                Q(assigned_to_students=student) | 
                Q(assigned_to_grades__students=student)
            ).distinct().count(),
            'submitted_assignments': submissions.count(),
            'pending_assignments': Assignment.objects.filter(
                Q(assigned_to_students=student) | 
                Q(assigned_to_grades__students=student)
            ).exclude(submissions__student=student).distinct().count(),
            'late_submissions': submissions.filter(is_late=True).count(),
            'average_grade': submissions.filter(grade__isnull=False).aggregate(
                avg=Avg('grade__marks_obtained')
            )['avg'] or 0,
        }
        
    else:
        # Admin statistics
        stats = {
            'total_assignments': Assignment.objects.count(),
            'published_assignments': Assignment.objects.filter(status='PUBLISHED').count(),
            'draft_assignments': Assignment.objects.filter(status='DRAFT').count(),
            'overdue_assignments': Assignment.objects.filter(
                status='PUBLISHED',
                due_date__lt=timezone.now()
            ).count(),
            'total_submissions': AssignmentSubmission.objects.count(),
            'graded_submissions': AssignmentSubmission.objects.filter(grade__isnull=False).count(),
            'pending_grades': AssignmentSubmission.objects.filter(grade__isnull=True).count(),
            'average_grade': AssignmentSubmission.objects.filter(grade__isnull=False).aggregate(
                avg=Avg('grade__marks_obtained')
            )['avg'] or 0,
        }
    
    return render(request, 'dashboard/assignments/statistics.html', {'stats': stats})


# ==================== AJAX AND FILE UPLOAD HANDLERS ====================

@login_required
def assignment_file_upload(request):
    """Handle file uploads for assignments"""
    if request.method == 'POST':
        try:
            assignment_id = request.POST.get('assignment_id')
            file = request.FILES.get('file')
            description = request.POST.get('description', '')
            
            if not assignment_id or not file:
                return JsonResponse({'success': False, 'error': 'Missing required fields'})
            
            assignment = get_object_or_404(Assignment, id=assignment_id)
            
            # Check permissions
            if hasattr(request.user, 'faculty_profile') and assignment.faculty != request.user.faculty_profile:
                return JsonResponse({'success': False, 'error': 'Permission denied'})
            
            # Create file record
            assignment_file = AssignmentFile.objects.create(
                assignment=assignment,
                file_name=file.name,
                file_path=file,
                file_type='ASSIGNMENT',
                uploaded_by=request.user,
                description=description
            )
            
            return JsonResponse({
                'success': True,
                'file_id': str(assignment_file.id),
                'file_name': assignment_file.file_name,
                'file_url': assignment_file.file_path.url,
                'file_size': assignment_file.file_size
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def assignment_publish_ajax(request, assignment_id):
    """AJAX endpoint to publish an assignment"""
    if request.method == 'POST':
        try:
            assignment = get_object_or_404(Assignment, id=assignment_id)
            
            if not hasattr(request.user, 'faculty_profile') or assignment.faculty != request.user.faculty_profile:
                return JsonResponse({'success': False, 'error': 'Permission denied'})
            
            if assignment.status != 'DRAFT':
                return JsonResponse({'success': False, 'error': 'Only draft assignments can be published'})
            
            assignment.status = 'PUBLISHED'
            assignment.save()
            
            return JsonResponse({'success': True, 'message': 'Assignment published successfully'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def assignment_close_ajax(request, assignment_id):
    """AJAX endpoint to close an assignment"""
    if request.method == 'POST':
        try:
            assignment = get_object_or_404(Assignment, id=assignment_id)
            
            if not hasattr(request.user, 'faculty_profile') or assignment.faculty != request.user.faculty_profile:
                return JsonResponse({'success': False, 'error': 'Permission denied'})
            
            if assignment.status not in ['PUBLISHED', 'DRAFT']:
                return JsonResponse({'success': False, 'error': 'Only published or draft assignments can be closed'})
            
            assignment.status = 'CLOSED'
            assignment.save()
            
            return JsonResponse({'success': True, 'message': 'Assignment closed successfully'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def assignment_comment_ajax(request, assignment_id):
    """AJAX endpoint to add comments to assignments"""
    if request.method == 'POST':
        try:
            assignment = get_object_or_404(Assignment, id=assignment_id)
            content = request.POST.get('content', '').strip()
            comment_type = request.POST.get('comment_type', 'GENERAL')
            
            if not content:
                return JsonResponse({'success': False, 'error': 'Comment content is required'})
            
            comment = AssignmentComment.objects.create(
                assignment=assignment,
                author=request.user,
                content=content,
                comment_type=comment_type
            )
            
            return JsonResponse({
                'success': True,
                'comment_id': str(comment.id),
                'author': request.user.email,
                'content': comment.content,
                'comment_type': comment.get_comment_type_display(),
                'created_at': comment.created_at.strftime('%M %d, %Y %H:%i')
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def assignment_stats_ajax(request):
    """AJAX endpoint to get assignment statistics"""
    try:
        user = request.user
        
        if hasattr(user, 'faculty_profile'):
            faculty = user.faculty_profile
            assignments = Assignment.objects.filter(faculty=faculty)
            submissions = AssignmentSubmission.objects.filter(assignment__faculty=faculty)
            
            stats = {
                'total_assignments': assignments.count(),
                'published_assignments': assignments.filter(status='PUBLISHED').count(),
                'draft_assignments': assignments.filter(status='DRAFT').count(),
                'overdue_assignments': assignments.filter(
                    status='PUBLISHED',
                    due_date__lt=timezone.now()
                ).count(),
                'total_submissions': submissions.count(),
                'graded_submissions': submissions.filter(grade__isnull=False).count(),
                'pending_grades': submissions.filter(grade__isnull=True).count(),
                'average_grade': submissions.filter(grade__isnull=False).aggregate(
                    avg=Avg('grade__marks_obtained')
                )['avg'] or 0,
            }
            
        elif hasattr(user, 'student_profile'):
            student = user.student_profile
            submissions = AssignmentSubmission.objects.filter(student=student)
            
            stats = {
                'total_assignments': Assignment.objects.filter(
                    Q(assigned_to_students=student) | 
                    Q(assigned_to_grades__students=student)
                ).distinct().count(),
                'submitted_assignments': submissions.count(),
                'pending_assignments': Assignment.objects.filter(
                    Q(assigned_to_students=student) | 
                    Q(assigned_to_grades__students=student)
                ).exclude(submissions__student=student).distinct().count(),
                'late_submissions': submissions.filter(is_late=True).count(),
                'average_grade': submissions.filter(grade__isnull=False).aggregate(
                    avg=Avg('grade__marks_obtained')
                )['avg'] or 0,
            }
            
        else:
            stats = {
                'total_assignments': Assignment.objects.count(),
                'published_assignments': Assignment.objects.filter(status='PUBLISHED').count(),
                'draft_assignments': Assignment.objects.filter(status='DRAFT').count(),
                'overdue_assignments': Assignment.objects.filter(
                    status='PUBLISHED',
                    due_date__lt=timezone.now()
                ).count(),
                'total_submissions': AssignmentSubmission.objects.count(),
                'graded_submissions': AssignmentSubmission.objects.filter(grade__isnull=False).count(),
                'pending_grades': AssignmentSubmission.objects.filter(grade__isnull=True).count(),
                'average_grade': AssignmentSubmission.objects.filter(grade__isnull=False).aggregate(
                    avg=Avg('grade__marks_obtained')
                )['avg'] or 0,
            }
        
        return JsonResponse({'success': True, 'stats': stats})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def assignment_autocomplete(request):
    """AJAX endpoint for assignment autocomplete"""
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse({'results': []})
    
    user = request.user
    
    if hasattr(user, 'faculty_profile'):
        assignments = Assignment.objects.filter(
            faculty=user.faculty_profile,
            title__icontains=query
        )[:10]
    elif hasattr(user, 'student_profile'):
        student = user.student_profile
        assignments = Assignment.objects.filter(
            Q(assigned_to_students=student) | 
            Q(assigned_to_grades__students=student),
            title__icontains=query
        ).distinct()[:10]
    else:
        assignments = Assignment.objects.filter(title__icontains=query)[:10]
    
    results = []
    for assignment in assignments:
        results.append({
            'id': str(assignment.id),
            'title': assignment.title,
            'faculty': assignment.faculty.name,
            'due_date': assignment.due_date.strftime('%M %d, %Y'),
            'status': assignment.get_status_display()
        })
    
    return JsonResponse({'results': results})


@login_required
def assignment_bulk_action(request):
    """AJAX endpoint for bulk actions on assignments"""
    if request.method == 'POST':
        try:
            action = request.POST.get('action')
            assignment_ids = request.POST.get('assignment_ids', '').split(',')
            
            if not action or not assignment_ids:
                return JsonResponse({'success': False, 'error': 'Missing required parameters'})
            
            user = request.user
            
            if not hasattr(user, 'faculty_profile'):
                return JsonResponse({'success': False, 'error': 'Permission denied'})
            
            assignments = Assignment.objects.filter(
                id__in=assignment_ids,
                faculty=user.faculty_profile
            )
            
            if action == 'publish':
                assignments.filter(status='DRAFT').update(status='PUBLISHED')
                message = f'{assignments.filter(status="PUBLISHED").count()} assignments published'
            elif action == 'close':
                assignments.filter(status__in=['PUBLISHED', 'DRAFT']).update(status='CLOSED')
                message = f'{assignments.filter(status="CLOSED").count()} assignments closed'
            elif action == 'delete':
                count = assignments.count()
                assignments.delete()
                message = f'{count} assignments deleted'
            else:
                return JsonResponse({'success': False, 'error': 'Invalid action'})
            
            return JsonResponse({'success': True, 'message': message})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def filter_students_ajax(request):
    """AJAX endpoint to filter students based on department, year, and section"""
    if request.method == 'GET':
        department_id = request.GET.get('department_id')
        academic_year = request.GET.get('academic_year')
        semester = request.GET.get('semester')
        course_section_id = request.GET.get('course_section_id')
        
        try:
            # Start with all students
            students = Student.objects.all()
            
            # Filter by department (through course enrollments)
            if department_id:
                from academics.models import CourseEnrollment
                student_ids = CourseEnrollment.objects.filter(
                    course_section__course__department_id=department_id,
                    status='ENROLLED'
                ).values_list('student_id', flat=True)
                students = students.filter(id__in=student_ids)
            
            # Filter by academic year
            if academic_year:
                students = students.filter(academic_year=academic_year)
            
            # Filter by course section
            if course_section_id:
                from academics.models import CourseEnrollment
                student_ids = CourseEnrollment.objects.filter(
                    course_section_id=course_section_id,
                    status='ENROLLED'
                ).values_list('student_id', flat=True)
                students = students.filter(id__in=student_ids)
            
            # Limit results for performance
            students = students[:100]
            
            # Format response
            student_data = []
            for student in students:
                student_data.append({
                    'id': str(student.id),
                    'name': f"{student.first_name} {student.last_name}",
                    'roll_number': student.roll_number,
                    'academic_year': student.academic_year or '',
                    'section': student.section or '',
                    'grade_level': student.grade_level
                })
            
            return JsonResponse({
                'success': True,
                'students': student_data,
                'count': len(student_data)
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


from django.contrib.auth.views import LoginView
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator

class CustomLoginView(LoginView):
    """Custom login view that handles email-based authentication"""
    template_name = 'dashboard/login.html'
    redirect_authenticated_user = True
    
    def get_success_url(self):
        return '/dashboard/'
    
    def form_valid(self, form):
        """Override to add custom success message"""
        from django.contrib import messages
        messages.success(self.request, f'Welcome back, {form.get_user().email}!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        """Override to add custom error message"""
        from django.contrib import messages
        messages.error(self.request, 'Invalid email or password.')
        return super().form_invalid(form)

from django.views.decorators.csrf import ensure_csrf_cookie, csrf_protect

@ensure_csrf_cookie
@csrf_protect
def custom_login(request):
    """Simple login view for testing"""
    from django.contrib.auth import authenticate, login
    from django.contrib import messages
    from django.shortcuts import render, redirect
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                if user.is_active:
                    login(request, user)
                    messages.success(request, f'Welcome back, {user.email}!')
                    return redirect('dashboard:home')
                else:
                    messages.error(request, 'Your account is disabled.')
            else:
                messages.error(request, 'Invalid email or password.')
        else:
            messages.error(request, 'Please enter both email and password.')
    
    # For GET requests, show the login form
    from django.contrib.auth.forms import AuthenticationForm
    form = AuthenticationForm()
    return render(request, 'dashboard/login.html', {'form': form})


# Student Division Management Views
@login_required
@user_passes_test(is_admin)
def student_divisions(request):
    """Student division management dashboard"""
    # Get filter parameters
    department_id = request.GET.get('department')
    academic_program_id = request.GET.get('academic_program')
    academic_year = request.GET.get('academic_year')
    year_of_study = request.GET.get('year_of_study')
    semester = request.GET.get('semester')
    section = request.GET.get('section')
    
    # Get all departments and programs for filters
    departments = Department.objects.filter(is_active=True).order_by('name')
    programs = AcademicProgram.objects.filter(is_active=True).order_by('name')
    
    # Get students with filters
    students = Student.objects.filter(status='ACTIVE')
    
    if department_id:
        students = students.filter(department_id=department_id)
    if academic_program_id:
        students = students.filter(academic_program_id=academic_program_id)
    if academic_year:
        students = students.filter(academic_year=academic_year)
    if year_of_study:
        students = students.filter(year_of_study=year_of_study)
    if semester:
        students = students.filter(semester=semester)
    if section:
        students = students.filter(section=section)
    
    # Group students by division
    divisions = {}
    for student in students:
        dept_code = student.department.code if student.department else 'No Department'
        program_code = student.academic_program.code if student.academic_program else 'No Program'
        
        if dept_code not in divisions:
            divisions[dept_code] = {
                'department': student.department,
                'programs': {}
            }
        
        if program_code not in divisions[dept_code]['programs']:
            divisions[dept_code]['programs'][program_code] = {
                'program': student.academic_program,
                'years': {}
            }
        
        year_key = f"{student.academic_year or 'No Year'}"
        if year_key not in divisions[dept_code]['programs'][program_code]['years']:
            divisions[dept_code]['programs'][program_code]['years'][year_key] = {
                'academic_year': student.academic_year,
                'year_of_study': {},
                'total_students': 0
            }
        
        study_year = student.year_of_study or 'No Year'
        if study_year not in divisions[dept_code]['programs'][program_code]['years'][year_key]['year_of_study']:
            divisions[dept_code]['programs'][program_code]['years'][year_key]['year_of_study'][study_year] = {
                'semesters': {},
                'total_students': 0
            }
        
        sem = student.semester or 'No Semester'
        if sem not in divisions[dept_code]['programs'][program_code]['years'][year_key]['year_of_study'][study_year]['semesters']:
            divisions[dept_code]['programs'][program_code]['years'][year_key]['year_of_study'][study_year]['semesters'][sem] = {
                'sections': {},
                'total_students': 0
            }
        
        sec = student.section or 'No Section'
        if sec not in divisions[dept_code]['programs'][program_code]['years'][year_key]['year_of_study'][study_year]['semesters'][sem]['sections']:
            divisions[dept_code]['programs'][program_code]['years'][year_key]['year_of_study'][study_year]['semesters'][sem]['sections'][sec] = {
                'students': [],
                'count': 0
            }
        
        divisions[dept_code]['programs'][program_code]['years'][year_key]['year_of_study'][study_year]['semesters'][sem]['sections'][sec]['students'].append(student)
        divisions[dept_code]['programs'][program_code]['years'][year_key]['year_of_study'][study_year]['semesters'][sem]['sections'][sec]['count'] += 1
        divisions[dept_code]['programs'][program_code]['years'][year_key]['year_of_study'][study_year]['semesters'][sem]['total_students'] += 1
        divisions[dept_code]['programs'][program_code]['years'][year_key]['year_of_study'][study_year]['total_students'] += 1
        divisions[dept_code]['programs'][program_code]['years'][year_key]['total_students'] += 1
    
    context = {
        'divisions': divisions,
        'departments': departments,
        'programs': programs,
        'year_choices': Student.YEAR_OF_STUDY_CHOICES,
        'semester_choices': Student.SEMESTER_CHOICES,
        'section_choices': Student.SECTION_CHOICES,
        'filters': {
            'department_id': department_id,
            'academic_program_id': academic_program_id,
            'academic_year': academic_year,
            'year_of_study': year_of_study,
            'semester': semester,
            'section': section,
        }
    }
    
    return render(request, 'dashboard/student_divisions.html', context)


@login_required
@user_passes_test(is_admin)
def student_assignments(request):
    """Student assignment management page"""
    if request.method == 'POST':
        # Handle student assignment
        student_ids = request.POST.getlist('student_ids')
        department_id = request.POST.get('department_id')
        academic_program_id = request.POST.get('academic_program_id')
        academic_year = request.POST.get('academic_year')
        year_of_study = request.POST.get('year_of_study')
        semester = request.POST.get('semester')
        section = request.POST.get('section')
        
        try:
            updated_count = 0
            for student_id in student_ids:
                student = Student.objects.get(id=student_id)
                
                if department_id:
                    student.department_id = department_id
                if academic_program_id:
                    student.academic_program_id = academic_program_id
                if academic_year:
                    student.academic_year = academic_year
                if year_of_study:
                    student.year_of_study = year_of_study
                if semester:
                    student.semester = semester
                if section:
                    student.section = section
                
                student.updated_by = request.user
                student.save()
                updated_count += 1
            
            messages.success(request, f'Successfully updated {updated_count} students.')
            return redirect('dashboard:student_assignments')
            
        except Exception as e:
            messages.error(request, f'Error updating students: {str(e)}')
    
    # Get all departments and programs
    departments = Department.objects.filter(is_active=True).order_by('name')
    programs = AcademicProgram.objects.filter(is_active=True).order_by('name')
    
    # Get students for assignment
    students = Student.objects.filter(status='ACTIVE').order_by('roll_number')
    
    context = {
        'students': students,
        'departments': departments,
        'programs': programs,
        'year_choices': Student.YEAR_OF_STUDY_CHOICES,
        'semester_choices': Student.SEMESTER_CHOICES,
        'section_choices': Student.SECTION_CHOICES,
    }
    
    return render(request, 'dashboard/student_assignments.html', context)


@login_required
@user_passes_test(is_admin)
def student_division_statistics(request):
    """Student division statistics page"""
    # Get filter parameters
    department_id = request.GET.get('department')
    academic_year = request.GET.get('academic_year')
    
    # Get all departments
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    # Get students with filters
    students = Student.objects.filter(status='ACTIVE')
    
    if department_id:
        students = students.filter(department_id=department_id)
    if academic_year:
        students = students.filter(academic_year=academic_year)
    
    # Calculate statistics
    stats = {}
    for dept in departments:
        dept_students = students.filter(department=dept)
        if dept_students.exists():
            # Count by year of study
            year_counts = {}
            for year, _ in Student.YEAR_OF_STUDY_CHOICES:
                count = dept_students.filter(year_of_study=year).count()
                if count > 0:
                    year_counts[year] = count
            
            # Count by semester
            semester_counts = {}
            for semester, _ in Student.SEMESTER_CHOICES:
                count = dept_students.filter(semester=semester).count()
                if count > 0:
                    semester_counts[semester] = count
            
            # Count by gender
            gender_counts = {}
            for gender, _ in Student.GENDER_CHOICES:
                count = dept_students.filter(gender=gender).count()
                if count > 0:
                    gender_counts[gender] = count
            
            # Count by program
            program_counts = {}
            for program in AcademicProgram.objects.filter(department=dept, is_active=True):
                count = dept_students.filter(academic_program=program).count()
                if count > 0:
                    program_counts[program.code] = {
                        'name': program.name,
                        'count': count
                    }
            
            stats[dept.code] = {
                'department': dept,
                'total_students': dept_students.count(),
                'by_year_of_study': year_counts,
                'by_semester': semester_counts,
                'by_gender': gender_counts,
                'by_program': program_counts
            }
    
    context = {
        'stats': stats,
        'departments': departments,
        'year_choices': Student.YEAR_OF_STUDY_CHOICES,
        'semester_choices': Student.SEMESTER_CHOICES,
        'filters': {
            'department_id': department_id,
            'academic_year': academic_year,
        }
    }
    
    return render(request, 'dashboard/student_division_statistics.html', context)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def bulk_assign_students(request):
    """API endpoint for bulk student assignment"""
    try:
        data = request.data
        student_ids = data.get('student_ids', [])
        assignment_data = data.get('assignment', {})
        
        updated_count = 0
        errors = []
        
        for student_id in student_ids:
            try:
                student = Student.objects.get(id=student_id)
                
                if assignment_data.get('department_id'):
                    student.department_id = assignment_data['department_id']
                if assignment_data.get('academic_program_id'):
                    student.academic_program_id = assignment_data['academic_program_id']
                if assignment_data.get('academic_year'):
                    student.academic_year = assignment_data['academic_year']
                if assignment_data.get('year_of_study'):
                    student.year_of_study = assignment_data['year_of_study']
                if assignment_data.get('semester'):
                    student.semester = assignment_data['semester']
                if assignment_data.get('section'):
                    student.section = assignment_data['section']
                
                student.updated_by = request.user
                student.save()
                updated_count += 1
                
            except Student.DoesNotExist:
                errors.append(f'Student with ID {student_id} not found')
            except Exception as e:
                errors.append(f'Error updating student {student_id}: {str(e)}')
        
        return Response({
            'success': True,
            'updated_count': updated_count,
            'errors': errors
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=400)